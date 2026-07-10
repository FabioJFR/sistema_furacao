import csv
import io
import logging
import os

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.utils.translation import gettext as _
from django.utils import timezone
from datetime import timedelta
from django.db.models import Count, Sum
from django.db.models.functions import TruncMonth
from django.db.utils import OperationalError, ProgrammingError
from django.db import transaction
from openpyxl import Workbook

from core.permissions import admin_required
from projetos.forms import (
    AcaoPreventivaForm,
    AcaoCorretivaForm,
    AgendamentoRelatorioExecutivoForm,
    AuditoriaHSEForm,
    ChecklistHSEForm,
    EvidenciaComplianceForm,
    FechoAcaoCorretivaForm,
    FornecedorCompraForm,
    IncidenteSegurancaForm,
    NotificacaoGestaoForm,
    PedidoCompraForm,
    PlanoAuditoriaHSEForm,
    PropostaFornecedorCompraForm,
    RelatorioExecutivoEmailForm,
)
from projetos.models import (
    AcaoPreventiva,
    AcaoCorretiva,
    AssiduidadeRegisto,
    AgendamentoRelatorioExecutivo,
    AuditoriaHSE,
    ChecklistHSE,
    Despesa,
    Empregados,
    EvidenciaCompliance,
    FechoAcaoCorretiva,
    FornecedorCompra,
    Furo,
    IncidenteSeguranca,
    MaquinaAvaria,
    NotificacaoGestao,
    PedidoCompra,
    PlaneamentoTurno,
    PlanoAuditoriaHSE,
    PropostaFornecedorCompra,
    Projeto,
    RegistoDiarioEmpregado,
    HistoricoEnvioRelatorioExecutivo,
)
from projetos.services.acesso_contexto import obter_empresa_admin_contexto
from projetos.services.gestao_compliance import (
    construir_dashboard_eficacia_compliance,
    gerar_dashboard_compliance_csv_bytes,
    gerar_auditorias_recorrentes_pendentes,
    guardar_evidencia_compliance_form,
    normalizar_filtros_dashboard_compliance,
    registar_fecho_formal_acao_corretiva,
    sincronizar_alertas_automaticos_compliance,
)
from projetos.services.gestao_compras import (
    avaliar_propostas_pedido,
    filtrar_pedidos_compra,
    normalizar_filtros_compras,
)
from projetos.services.gestao_relatorios import (
    calcular_proximo_envio_agendado,
    construir_filtros_periodo_agendamento,
    construir_url_relatorio_com_filtros,
    enviar_relatorio_executivo_email,
    normalizar_filtros_relatorio_executivo,
    normalizar_destinos,
    resolver_destinos_relatorio,
)

logger = logging.getLogger("core")


def _resolver_empresa_admin(request):
    empresa, resposta_erro = obter_empresa_admin_contexto(
        request=request,
        mensagem_sem_permissao="Não tens permissão para aceder a esta área.",
        mensagem_sem_empresa="O utilizador administrador não está associado a uma empresa válida.",
        redirect_sem_permissao="projetos:redirect_after_login",
        redirect_sem_empresa="projetos:dashboard",
    )
    if resposta_erro:
        return None, resposta_erro
    return empresa, None


def _obter_ou_criar_agendamento_relatorio(empresa):
    agendamento, _ = AgendamentoRelatorioExecutivo.objects.get_or_create(
        empresa=empresa,
        defaults={
            "ativo": False,
            "frequencia": "semanal",
            "dia_semana": 0,
            "dia_mes": 1,
            "incluir_csv": True,
            "incluir_xlsx": True,
            "incluir_pdf": False,
        },
    )
    return agendamento


def _resumo_kpis_relatorio(relatorio):
    return {
        "projetos": relatorio["kpis"][0]["valor"],
        "furos": relatorio["kpis"][1]["valor"],
        "empregados": relatorio["kpis"][2]["valor"],
        "despesa_periodo": relatorio["kpis"][3]["valor"],
    }


def _registar_historico_envio_relatorio(
    *,
    empresa,
    agendamento=None,
    origem="manual",
    status="sucesso",
    assunto="",
    destinos=None,
    incluir_csv=True,
    incluir_xlsx=True,
    incluir_pdf=False,
    enviados=0,
    filtros=None,
    relatorio=None,
    erro="",
):
    try:
        HistoricoEnvioRelatorioExecutivo.objects.create(
            empresa=empresa,
            agendamento=agendamento,
            origem=origem,
            status=status,
            assunto=(assunto or "").strip(),
            destinos="\n".join(destinos or []),
            incluir_csv=bool(incluir_csv),
            incluir_xlsx=bool(incluir_xlsx),
            incluir_pdf=bool(incluir_pdf),
            enviados=max(int(enviados or 0), 0),
            filtros_json=filtros or {},
            resumo_json=_resumo_kpis_relatorio(relatorio) if relatorio else {},
            erro=(erro or "").strip(),
        )
    except (ProgrammingError, OperationalError):
        logger.warning(
            "Histórico de relatórios não registado por falta de migração. empresa_id=%s origem=%s",
            getattr(empresa, "id", None),
            origem,
        )


def _render_secao(request, *, slug, titulo, descricao, proximos_passos, kpis=None, linhas=None):
    return render(
        request,
        "projetos/gestao_secao.html",
        {
            "slug": slug,
            "titulo": titulo,
            "descricao": descricao,
            "proximos_passos": proximos_passos,
            "kpis": kpis or [],
            "linhas": linhas or [],
        },
    )


def _render_formulario_evidencia_compliance(request, *, form, origem, origem_label):
    return render(
        request,
        "projetos/gestao_evidencia_compliance_form.html",
        {
            "form": form,
            "origem": origem,
            "origem_label": origem_label,
        },
    )


def _criar_evidencia_compliance_para_origem(request, *, origem, origem_label, empresa, kwargs_origem):
    form = EvidenciaComplianceForm(request.POST or None, request.FILES or None)
    if request.method == "POST" and form.is_valid():
        guardar_evidencia_compliance_form(
            form=form,
            empresa=empresa,
            user=request.user,
            **kwargs_origem,
        )
        messages.success(request, _("Evidência adicionada com sucesso."))
        return redirect("projetos:gestao_compliance_seguranca")
    return _render_formulario_evidencia_compliance(
        request,
        form=form,
        origem=origem,
        origem_label=origem_label,
    )


def _apagar_evidencia_compliance(request, *, evidencia):
    if request.method == "POST":
        evidencia.delete()
        messages.success(request, _("Evidência apagada com sucesso."))
        return redirect("projetos:gestao_compliance_seguranca")
    return render(
        request,
        "projetos/gestao_evidencia_compliance_confirm_delete.html",
        {"item": evidencia},
    )


@login_required
@admin_required
def gestao_hub(request):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro

    agora = timezone.now()
    sla_vencido = NotificacaoGestao.objects.filter(
        empresa=empresa,
        estado__in=["aberta", "em_andamento"],
        prazo__lt=agora,
    ).count()
    pedidos_pendentes = PedidoCompra.objects.filter(empresa=empresa, estado="pendente").count()
    avarias_abertas = MaquinaAvaria.objects.filter(empresa=empresa).exclude(status="resolvida").count()

    secoes = [
        {
            "titulo": _("Clientes & Contratos"),
            "descricao": _("Gestão de clientes, condições comerciais, SLA e documentos contratuais."),
            "url_name": "projetos:cliente_contrato_list",
            "icone": "🤝",
        },
        {
            "titulo": _("Planeamento"),
            "descricao": _("Planeamento de equipas, turnos, furos e máquinas com deteção de conflitos."),
            "url_name": "projetos:gestao_planeamento",
            "icone": "🗓️",
        },
        {
            "titulo": _("RH & Assiduidade"),
            "descricao": _("Férias, faltas, horas extra, banco de horas e validação de presença."),
            "url_name": "projetos:gestao_rh_assiduidade",
            "icone": "👥",
        },
        {
            "titulo": _("Compras & Fornecedores"),
            "descricao": _("Pedidos de compra, fornecedores, comparação de preço e histórico."),
            "url_name": "projetos:gestao_compras_fornecedores",
            "icone": "🛒",
        },
        {
            "titulo": _("Compliance & Segurança"),
            "descricao": _("Checklist HSE, incidentes, auditorias e validade documental."),
            "url_name": "projetos:gestao_compliance_seguranca",
            "icone": "🛡️",
        },
        {
            "titulo": _("Centro de Notificações"),
            "descricao": _("Fila de alertas e ações pendentes por prioridade operacional."),
            "url_name": "projetos:gestao_notificacoes",
            "icone": "🔔",
        },
        {
            "titulo": _("Relatórios Executivos"),
            "descricao": _("KPIs executivos semanais/mensais com exportação para gestão."),
            "url_name": "projetos:gestao_relatorios_executivos",
            "icone": "📑",
        },
    ]
    return render(
        request,
        "projetos/gestao_hub.html",
        {
            "secoes": secoes,
            "kpis": [
                {"titulo": _("SLA vencido"), "valor": str(sla_vencido)},
                {"titulo": _("Pedidos pendentes"), "valor": str(pedidos_pendentes)},
                {"titulo": _("Avarias em aberto"), "valor": str(avarias_abertas)},
            ],
        },
    )


@login_required
@admin_required
def gestao_clientes_contratos(request):
    return redirect("projetos:cliente_contrato_list")


@login_required
@admin_required
def gestao_planeamento(request):
    return redirect("projetos:planeamento_turno_list")


@login_required
@admin_required
def gestao_rh_assiduidade(request):
    return redirect("projetos:assiduidade_list")


@login_required
@admin_required
def gestao_compras_fornecedores(request):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro

    hoje = timezone.localdate()
    filtros = normalizar_filtros_compras(request.GET)
    despesas = Despesa.objects.filter(empresa=empresa)
    total = despesas.aggregate(total=Sum("valor"))["total"] or 0
    total_mes = despesas.filter(data__year=hoje.year, data__month=hoje.month).aggregate(total=Sum("valor"))["total"] or 0
    categorias = (
        despesas.values("categoria")
        .annotate(total=Sum("valor"), qtd=Count("id"))
        .order_by("-total")[:8]
    )
    recentes = despesas.select_related("projeto", "furo", "maquina").order_by("-data", "-criado_em")[:8]
    pedidos_qs = filtrar_pedidos_compra(empresa=empresa, filtros=filtros)
    pedidos = pedidos_qs[:20]
    pedidos_info = []
    projetos_choices = Projeto.objects.filter(empresa=empresa).order_by("nome").values("id", "nome")
    kpi_pedidos_total = PedidoCompra.objects.filter(empresa=empresa).count()
    kpi_pedidos_filtrados = pedidos_qs.count()
    fornecedores_top = (
        PedidoCompra.objects.filter(empresa=empresa)
        .exclude(fornecedor_sugerido__exact="")
        .values("fornecedor_sugerido")
        .annotate(
            qtd=Count("id"),
            total_estimado=Sum("valor_estimado"),
        )
        .order_by("-qtd", "-total_estimado")[:8]
    )
    fornecedores_ativos = (
        PedidoCompra.objects.filter(empresa=empresa)
        .exclude(fornecedor_sugerido__exact="")
        .values("fornecedor_sugerido")
        .distinct()
        .count()
    )
    fornecedores_cadastrados = []
    pendentes_por_prioridade_qs = (
        PedidoCompra.objects.filter(empresa=empresa, estado="pendente")
        .values("prioridade")
        .annotate(qtd=Count("id"))
        .order_by("-qtd")
    )
    pendentes_label_map = dict(PedidoCompra.PRIORIDADE_CHOICES)
    pendentes_por_prioridade = [
        {
            "prioridade": item["prioridade"],
            "prioridade_label": pendentes_label_map.get(item["prioridade"], item["prioridade"] or "-"),
            "qtd": item["qtd"],
        }
        for item in pendentes_por_prioridade_qs
    ]

    try:
        for pedido in pedidos:
            propostas_qs = pedido.propostas_fornecedor.all().order_by("valor_proposto", "prazo_entrega_dias")
            proposta_selecionada = propostas_qs.filter(selecionada=True).first()
            melhor_proposta = propostas_qs.first()
            pedidos_info.append(
                {
                    "pedido": pedido,
                    "propostas_count": propostas_qs.count(),
                    "proposta_selecionada": proposta_selecionada,
                    "melhor_proposta": melhor_proposta,
                }
            )
        fornecedores_cadastrados = FornecedorCompra.objects.filter(empresa=empresa).order_by("nome")
    except (ProgrammingError, OperationalError):
        messages.warning(
            request,
            _(
                "Módulo de fornecedores/propostas ainda sem migração aplicada. "
                "Executa `python manage.py migrate` para ativar esta funcionalidade."
            ),
        )
        pedidos_info = [{"pedido": pedido, "propostas_count": 0, "proposta_selecionada": None, "melhor_proposta": None} for pedido in pedidos]
        fornecedores_cadastrados = []

    return render(
        request,
        "projetos/gestao_compras_fornecedores.html",
        {
            "titulo": _("Compras & Fornecedores"),
            "descricao": _("Resumo financeiro de compras e despesas com foco operacional."),
            "kpis": [
                {"titulo": _("Despesa total"), "valor": f"{total:,.2f} €"},
                {"titulo": _("Despesa do mês"), "valor": f"{total_mes:,.2f} €"},
                {"titulo": _("Registos"), "valor": str(despesas.count())},
                {"titulo": _("Fornecedores ativos"), "valor": str(fornecedores_ativos)},
            ],
            "categorias": categorias,
            "pedidos_info": pedidos_info,
            "filtros": filtros,
            "projetos_choices": projetos_choices,
            "estado_choices": [("", _("Todos"))] + list(PedidoCompra.ESTADO_CHOICES),
            "prioridade_choices": [("", _("Todos"))] + list(PedidoCompra.PRIORIDADE_CHOICES),
            "kpi_pedidos_total": kpi_pedidos_total,
            "kpi_pedidos_filtrados": kpi_pedidos_filtrados,
            "despesas_recentes": recentes,
            "fornecedores_top": fornecedores_top,
            "fornecedores_cadastrados": fornecedores_cadastrados,
            "pendentes_por_prioridade": pendentes_por_prioridade,
        },
    )


@login_required
@admin_required
def gestao_compras_export_csv(request):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro

    filtros = normalizar_filtros_compras(request.GET)
    pedidos = filtrar_pedidos_compra(empresa=empresa, filtros=filtros)

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="pedidos_compra.csv"'
    writer = csv.writer(response)
    writer.writerow(
        [
            "Data",
            "Descricao",
            "Projeto",
            "Solicitado por",
            "Categoria",
            "Fornecedor",
            "Prioridade",
            "Estado",
            "Valor estimado",
            "Data necessidade",
        ]
    )
    for p in pedidos:
        writer.writerow(
            [
                p.criado_em.strftime("%d/%m/%Y"),
                p.descricao,
                p.projeto.nome if p.projeto else "",
                p.solicitado_por.nome if p.solicitado_por else "",
                p.categoria or "",
                p.fornecedor_sugerido or "",
                p.get_prioridade_display(),
                p.get_estado_display(),
                f"{(p.valor_estimado or 0):.2f}",
                p.data_necessidade.strftime("%d/%m/%Y") if p.data_necessidade else "",
            ]
        )
    return response


@login_required
@admin_required
def gestao_compras_export_xlsx(request):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro

    filtros = normalizar_filtros_compras(request.GET)
    pedidos = filtrar_pedidos_compra(empresa=empresa, filtros=filtros)

    wb = Workbook()
    ws = wb.active
    ws.title = "Pedidos Compra"
    ws.append(
        [
            "Data",
            "Descricao",
            "Projeto",
            "Solicitado por",
            "Categoria",
            "Fornecedor",
            "Prioridade",
            "Estado",
            "Valor estimado",
            "Data necessidade",
        ]
    )
    for p in pedidos:
        ws.append(
            [
                p.criado_em.strftime("%d/%m/%Y"),
                p.descricao,
                p.projeto.nome if p.projeto else "",
                p.solicitado_por.nome if p.solicitado_por else "",
                p.categoria or "",
                p.fornecedor_sugerido or "",
                p.get_prioridade_display(),
                p.get_estado_display(),
                float(f"{(p.valor_estimado or 0):.2f}"),
                p.data_necessidade.strftime("%d/%m/%Y") if p.data_necessidade else "",
            ]
        )

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="pedidos_compra.xlsx"'
    wb.save(response)
    return response


@login_required
@admin_required
def gestao_pedido_compra_create(request):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    form = PedidoCompraForm(request.POST or None, empresa=empresa)
    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.empresa = empresa
        obj.save()
        messages.success(request, "Pedido de compra criado com sucesso.")
        return redirect("projetos:gestao_compras_fornecedores")
    return render(request, "projetos/gestao_pedido_compra_form.html", {"form": form, "is_create": True})


@login_required
@admin_required
def gestao_pedido_compra_update(request, pk):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    item = PedidoCompra.objects.filter(empresa=empresa, pk=pk).first()
    if not item:
        messages.error(request, "Pedido não encontrado.")
        return redirect("projetos:gestao_compras_fornecedores")
    form = PedidoCompraForm(request.POST or None, instance=item, empresa=empresa)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Pedido atualizado com sucesso.")
        return redirect("projetos:gestao_compras_fornecedores")
    return render(
        request,
        "projetos/gestao_pedido_compra_form.html",
        {"form": form, "is_create": False, "item": item},
    )


@login_required
@admin_required
def gestao_pedido_compra_delete(request, pk):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    item = PedidoCompra.objects.filter(empresa=empresa, pk=pk).first()
    if not item:
        messages.error(request, "Pedido não encontrado.")
        return redirect("projetos:gestao_compras_fornecedores")
    if request.method == "POST":
        item.delete()
        messages.success(request, "Pedido apagado com sucesso.")
        return redirect("projetos:gestao_compras_fornecedores")
    return render(request, "projetos/gestao_pedido_compra_confirm_delete.html", {"item": item})


@login_required
@admin_required
def gestao_pedido_compra_estado(request, pk, estado):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    if request.method != "POST":
        messages.error(request, "A atualização de estado deve ser feita por formulário.")
        return redirect("projetos:gestao_compras_fornecedores")
    item = PedidoCompra.objects.filter(empresa=empresa, pk=pk).first()
    if not item:
        messages.error(request, "Pedido não encontrado.")
        return redirect("projetos:gestao_compras_fornecedores")
    if estado not in {"aprovado", "rejeitado"}:
        messages.error(request, "Estado inválido.")
        return redirect("projetos:gestao_compras_fornecedores")
    item.estado = estado
    item.aprovado_em = timezone.now()
    item.save(update_fields=["estado", "aprovado_em", "atualizado_em"])
    messages.success(request, f"Pedido {item.get_estado_display().lower()} com sucesso.")
    return redirect("projetos:gestao_compras_fornecedores")


@login_required
@admin_required
def gestao_fornecedor_create(request):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    form = FornecedorCompraForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        fornecedor = form.save(commit=False)
        fornecedor.empresa = empresa
        fornecedor.save()
        messages.success(request, _("Fornecedor criado com sucesso."))
        return redirect("projetos:gestao_compras_fornecedores")
    return render(request, "projetos/gestao_fornecedor_form.html", {"form": form, "is_create": True})


@login_required
@admin_required
def gestao_fornecedor_update(request, pk):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    fornecedor = FornecedorCompra.objects.filter(empresa=empresa, pk=pk).first()
    if not fornecedor:
        messages.error(request, _("Fornecedor não encontrado."))
        return redirect("projetos:gestao_compras_fornecedores")
    form = FornecedorCompraForm(request.POST or None, instance=fornecedor)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, _("Fornecedor atualizado com sucesso."))
        return redirect("projetos:gestao_compras_fornecedores")
    return render(
        request,
        "projetos/gestao_fornecedor_form.html",
        {"form": form, "is_create": False, "item": fornecedor},
    )


@login_required
@admin_required
def gestao_fornecedor_delete(request, pk):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    fornecedor = FornecedorCompra.objects.filter(empresa=empresa, pk=pk).first()
    if not fornecedor:
        messages.error(request, _("Fornecedor não encontrado."))
        return redirect("projetos:gestao_compras_fornecedores")
    if request.method == "POST":
        fornecedor.delete()
        messages.success(request, _("Fornecedor apagado com sucesso."))
        return redirect("projetos:gestao_compras_fornecedores")
    return render(request, "projetos/gestao_fornecedor_confirm_delete.html", {"item": fornecedor})


@login_required
@admin_required
def gestao_proposta_compra_create(request, pedido_pk):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    pedido = PedidoCompra.objects.filter(empresa=empresa, pk=pedido_pk).first()
    if not pedido:
        messages.error(request, _("Pedido não encontrado."))
        return redirect("projetos:gestao_compras_fornecedores")
    form = PropostaFornecedorCompraForm(request.POST or None, empresa=empresa)
    if request.method == "POST" and form.is_valid():
        proposta = form.save(commit=False)
        proposta.pedido = pedido
        proposta.save()
        if proposta.selecionada:
            PropostaFornecedorCompra.objects.filter(pedido=pedido).exclude(pk=proposta.pk).update(selecionada=False)
        messages.success(request, _("Proposta registada com sucesso."))
        return redirect("projetos:gestao_pedido_compra_comparar", pk=pedido.pk)
    return render(
        request,
        "projetos/gestao_proposta_compra_form.html",
        {"form": form, "pedido": pedido, "is_create": True},
    )


@login_required
@admin_required
def gestao_proposta_compra_update(request, pk):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    proposta = PropostaFornecedorCompra.objects.select_related("pedido").filter(pedido__empresa=empresa, pk=pk).first()
    if not proposta:
        messages.error(request, _("Proposta não encontrada."))
        return redirect("projetos:gestao_compras_fornecedores")
    form = PropostaFornecedorCompraForm(request.POST or None, instance=proposta, empresa=empresa)
    if request.method == "POST" and form.is_valid():
        proposta = form.save()
        if proposta.selecionada:
            PropostaFornecedorCompra.objects.filter(pedido=proposta.pedido).exclude(pk=proposta.pk).update(selecionada=False)
        messages.success(request, _("Proposta atualizada com sucesso."))
        return redirect("projetos:gestao_pedido_compra_comparar", pk=proposta.pedido.pk)
    return render(
        request,
        "projetos/gestao_proposta_compra_form.html",
        {"form": form, "pedido": proposta.pedido, "is_create": False, "item": proposta},
    )


@login_required
@admin_required
def gestao_proposta_compra_delete(request, pk):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    proposta = PropostaFornecedorCompra.objects.select_related("pedido").filter(pedido__empresa=empresa, pk=pk).first()
    if not proposta:
        messages.error(request, _("Proposta não encontrada."))
        return redirect("projetos:gestao_compras_fornecedores")
    if request.method == "POST":
        pedido_pk = proposta.pedido.pk
        proposta.delete()
        messages.success(request, _("Proposta apagada com sucesso."))
        return redirect("projetos:gestao_pedido_compra_comparar", pk=pedido_pk)
    return render(request, "projetos/gestao_proposta_compra_confirm_delete.html", {"item": proposta})


@login_required
@admin_required
def gestao_pedido_compra_comparar(request, pk):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    pedido = PedidoCompra.objects.filter(empresa=empresa, pk=pk).first()
    if not pedido:
        messages.error(request, _("Pedido não encontrado."))
        return redirect("projetos:gestao_compras_fornecedores")

    propostas_avaliadas = avaliar_propostas_pedido(pedido=pedido)
    return render(
        request,
        "projetos/gestao_pedido_compra_comparar.html",
        {
            "pedido": pedido,
            "propostas_avaliadas": propostas_avaliadas,
        },
    )


@login_required
@admin_required
def gestao_pedido_compra_selecionar_melhor(request, pk):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    pedido = PedidoCompra.objects.filter(empresa=empresa, pk=pk).first()
    if not pedido:
        messages.error(request, _("Pedido não encontrado."))
        return redirect("projetos:gestao_compras_fornecedores")

    propostas_avaliadas = avaliar_propostas_pedido(pedido=pedido)
    if not propostas_avaliadas:
        messages.error(request, _("Não existem propostas para comparar neste pedido."))
        return redirect("projetos:gestao_pedido_compra_comparar", pk=pedido.pk)

    melhor = propostas_avaliadas[0]["obj"]
    with transaction.atomic():
        PropostaFornecedorCompra.objects.filter(pedido=pedido).update(selecionada=False)
        melhor.selecionada = True
        melhor.save(update_fields=["selecionada", "atualizado_em"])
    messages.success(request, _("Melhor proposta selecionada automaticamente com base em preço e prazo."))
    return redirect("projetos:gestao_pedido_compra_comparar", pk=pedido.pk)


@login_required
@admin_required
def gestao_pedido_compra_selecionar_proposta(request, pedido_pk, proposta_pk):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    pedido = PedidoCompra.objects.filter(empresa=empresa, pk=pedido_pk).first()
    if not pedido:
        messages.error(request, _("Pedido não encontrado."))
        return redirect("projetos:gestao_compras_fornecedores")
    proposta = PropostaFornecedorCompra.objects.filter(pedido=pedido, pk=proposta_pk).first()
    if not proposta:
        messages.error(request, _("Proposta não encontrada."))
        return redirect("projetos:gestao_pedido_compra_comparar", pk=pedido.pk)

    with transaction.atomic():
        PropostaFornecedorCompra.objects.filter(pedido=pedido).update(selecionada=False)
        proposta.selecionada = True
        proposta.save(update_fields=["selecionada", "atualizado_em"])
    messages.success(request, _("Proposta selecionada com sucesso."))
    return redirect("projetos:gestao_pedido_compra_comparar", pk=pedido.pk)


@login_required
@admin_required
def gestao_compliance_seguranca(request):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro

    hoje = timezone.localdate()
    filtros_dashboard = normalizar_filtros_dashboard_compliance(request.GET)
    avarias = MaquinaAvaria.objects.filter(empresa=empresa)
    avarias_abertas = avarias.exclude(status="resolvida").count()
    avarias_resolvidas_30d = avarias.filter(
        status="resolvida",
        atualizado_em__date__gte=hoje - timedelta(days=30),
    ).count()
    empregados = Empregados.objects.filter(empresa=empresa)
    sem_contrato = empregados.filter(contrato__isnull=True).count()
    sem_curriculo = empregados.filter(curriculo__isnull=True).count()
    filtro_check = filtros_dashboard["check_status"]
    filtro_inc = filtros_dashboard["inc_status"]
    filtro_aud = filtros_dashboard["aud_status"]
    filtro_acao = filtros_dashboard["acao_status"]
    filtro_plano = filtros_dashboard["plano_status"]
    filtro_prev = filtros_dashboard["prev_status"]
    drill_projeto = filtros_dashboard["drill_projeto"]
    drill_responsavel = filtros_dashboard["drill_responsavel"]
    janela_dias = filtros_dashboard["janela_dias"]
    checklists = []
    incidentes = []
    auditorias = []
    planos_auditoria = []
    acoes_corretivas = []
    acoes_preventivas = []
    checklists_nao_conformes = 0
    incidentes_abertos = 0
    auditorias_abertas = 0
    planos_vencidos = 0
    acoes_abertas = 0
    acoes_vencidas = 0
    acoes_preventivas_abertas = 0
    acoes_preventivas_vencidas = 0
    evidencias_total = 0
    acoes_com_fecho_formal = 0
    evidencias_recentes = []
    dashboard_eficacia = {
        "resumo": {
            "total_acoes": 0,
            "total_concluidas": 0,
            "total_abertas": 0,
            "total_vencidas": 0,
            "taxa_global_fecho": 0.0,
            "tempo_medio_fecho_global": 0.0,
            "taxa_sla_global": 0.0,
            "responsaveis_criticos": 0,
            "projetos_em_risco": 0,
            "equipas_em_risco": 0,
        },
        "responsaveis": [],
        "equipas": [],
        "projetos": [],
        "tendencia": [],
        "historico": [],
        "comparativos": {
            "periodo_atual": None,
            "periodo_anterior": None,
            "delta_abertas": 0,
            "delta_concluidas": 0,
            "delta_vencidas": 0,
            "delta_sla": 0.0,
            "delta_tempo_medio": 0.0,
        },
        "benchmark": {
            "melhor_responsavel": None,
            "responsavel_maior_risco": None,
            "melhor_equipa": None,
            "equipa_maior_risco": None,
        },
        "previsao_incumprimento": {
            "responsaveis": [],
            "projetos": [],
        },
        "drilldown": {
            "titulo": "",
            "projeto_id": "",
            "responsavel_id": "",
            "projeto_choices": [],
            "responsavel_choices": [],
            "resumo": {
                "total": 0,
                "abertas": 0,
                "concluidas": 0,
                "vencidas": 0,
                "vence_7d": 0,
                "taxa_sla": 0.0,
                "tempo_medio_fecho": 0.0,
            },
            "itens": [],
        },
        "snapshots": {
            "janela_dias": 0,
            "atual": None,
            "anterior": None,
            "delta_abertas": 0,
            "delta_concluidas": 0,
            "delta_vencidas": 0,
            "delta_sla": 0.0,
            "delta_tempo_medio": 0.0,
        },
        "janela_dias": 0,
        "inicio_janela": None,
        "referencia": hoje,
    }
    alertas_compliance = {
        "criticos": [],
        "preventivos": [],
    }

    try:
        checklists_qs = ChecklistHSE.objects.filter(empresa=empresa).select_related("projeto", "responsavel").annotate(
            total_evidencias=Count("evidencias")
        )
        if filtro_check:
            checklists_qs = checklists_qs.filter(status=filtro_check)
        checklists = list(checklists_qs.order_by("-data_check", "-criado_em")[:25])
        checklists_nao_conformes = ChecklistHSE.objects.filter(empresa=empresa, status="nao_conforme").count()

        incidentes_qs = IncidenteSeguranca.objects.filter(empresa=empresa).select_related(
            "projeto",
            "reportado_por",
            "responsavel",
        ).annotate(total_evidencias=Count("evidencias"))
        if filtro_inc:
            incidentes_qs = incidentes_qs.filter(status=filtro_inc)
        incidentes = list(incidentes_qs.order_by("status", "-data_incidente", "-criado_em")[:25])
        incidentes_abertos = IncidenteSeguranca.objects.filter(empresa=empresa).exclude(status="fechado").count()

        auditorias_qs = AuditoriaHSE.objects.filter(empresa=empresa).select_related("projeto", "responsavel").annotate(
            total_evidencias=Count("evidencias")
        )
        if filtro_aud:
            auditorias_qs = auditorias_qs.filter(status=filtro_aud)
        auditorias = list(auditorias_qs.order_by("status", "-data_auditoria", "-criado_em")[:25])
        auditorias_abertas = AuditoriaHSE.objects.filter(empresa=empresa).exclude(status="concluida").count()

        planos_qs = PlanoAuditoriaHSE.objects.filter(empresa=empresa).select_related("projeto", "responsavel")
        if filtro_plano == "ativos":
            planos_qs = planos_qs.filter(ativo=True)
        elif filtro_plano == "inativos":
            planos_qs = planos_qs.filter(ativo=False)
        elif filtro_plano == "vencidos":
            planos_qs = planos_qs.filter(ativo=True, proxima_execucao__lt=hoje)
        planos_auditoria = list(planos_qs.order_by("proxima_execucao", "titulo")[:25])
        planos_vencidos = PlanoAuditoriaHSE.objects.filter(
            empresa=empresa,
            ativo=True,
            proxima_execucao__lt=hoje,
        ).count()

        acoes_qs = AcaoCorretiva.objects.filter(empresa=empresa).select_related(
            "projeto",
            "responsavel",
            "checklist",
            "incidente",
            "auditoria",
            "fecho_formal",
        ).annotate(total_evidencias=Count("evidencias"))
        if filtro_acao:
            acoes_qs = acoes_qs.filter(status=filtro_acao)
        acoes_corretivas = list(acoes_qs.order_by("status", "prazo", "-criado_em")[:25])
        acoes_abertas = AcaoCorretiva.objects.filter(empresa=empresa).exclude(status__in=["concluida", "cancelada"]).count()
        acoes_vencidas = AcaoCorretiva.objects.filter(
            empresa=empresa,
            prazo__lt=hoje,
        ).exclude(status__in=["concluida", "cancelada"]).count()
        acoes_prev_qs = AcaoPreventiva.objects.filter(empresa=empresa).select_related(
            "projeto",
            "responsavel",
            "checklist",
            "incidente",
            "auditoria",
        ).annotate(total_evidencias=Count("evidencias"))
        if filtro_prev:
            acoes_prev_qs = acoes_prev_qs.filter(status=filtro_prev)
        acoes_preventivas = list(acoes_prev_qs.order_by("status", "prazo", "-criado_em")[:25])
        acoes_preventivas_abertas = AcaoPreventiva.objects.filter(empresa=empresa).exclude(
            status__in=["concluida", "cancelada"]
        ).count()
        acoes_preventivas_vencidas = AcaoPreventiva.objects.filter(
            empresa=empresa,
            prazo__lt=hoje,
        ).exclude(status__in=["concluida", "cancelada"]).count()
        evidencias_total = EvidenciaCompliance.objects.filter(empresa=empresa).count()
        acoes_com_fecho_formal = FechoAcaoCorretiva.objects.filter(empresa=empresa).count()
        evidencias_recentes = list(
            EvidenciaCompliance.objects.filter(empresa=empresa)
            .select_related("criado_por", "checklist", "incidente", "auditoria", "acao_corretiva", "acao_preventiva")
            .order_by("-criado_em")[:12]
        )
        dashboard_eficacia = construir_dashboard_eficacia_compliance(
            empresa=empresa,
            referencia=hoje,
            projeto_id=drill_projeto,
            responsavel_id=drill_responsavel,
            janela_dias=janela_dias or None,
        )
        alertas_compliance = sincronizar_alertas_automaticos_compliance(empresa=empresa, referencia=hoje)
    except (ProgrammingError, OperationalError):
        messages.warning(
            request,
            _("Módulo de Compliance ainda sem migração aplicada. Executa `python manage.py migrate` para ativar todos os dados."),
        )

    return render(
        request,
        "projetos/gestao_compliance_seguranca.html",
        {
            "titulo": _("Compliance & Segurança"),
            "descricao": _("Monitorização de risco operacional e conformidade documental."),
            "kpis": [
                {"titulo": _("Avarias em aberto"), "valor": str(avarias_abertas)},
                {"titulo": _("Avarias resolvidas (30d)"), "valor": str(avarias_resolvidas_30d)},
                {"titulo": _("Empregados sem contrato"), "valor": str(sem_contrato)},
                {"titulo": _("Empregados sem currículo"), "valor": str(sem_curriculo)},
                {"titulo": _("Checklists não conformes"), "valor": str(checklists_nao_conformes)},
                {"titulo": _("Incidentes abertos"), "valor": str(incidentes_abertos)},
                {"titulo": _("Auditorias abertas"), "valor": str(auditorias_abertas)},
                {"titulo": _("Planos vencidos"), "valor": str(planos_vencidos)},
                {"titulo": _("Ações corretivas abertas"), "valor": str(acoes_abertas)},
                {"titulo": _("Ações vencidas"), "valor": str(acoes_vencidas)},
                {"titulo": _("Ações preventivas abertas"), "valor": str(acoes_preventivas_abertas)},
                {"titulo": _("Preventivas vencidas"), "valor": str(acoes_preventivas_vencidas)},
                {"titulo": _("Evidências registadas"), "valor": str(evidencias_total)},
                {"titulo": _("Fechos formais"), "valor": str(acoes_com_fecho_formal)},
            ],
            "checklists": checklists,
            "incidentes": incidentes,
            "auditorias": auditorias,
            "planos_auditoria": planos_auditoria,
            "acoes_corretivas": acoes_corretivas,
            "acoes_preventivas": acoes_preventivas,
            "evidencias_recentes": evidencias_recentes,
            "dashboard_eficacia": dashboard_eficacia,
            "alertas_compliance": alertas_compliance,
            "check_status_choices": [("", _("Todos"))] + list(ChecklistHSE.STATUS_CHOICES),
            "inc_status_choices": [("", _("Todos"))] + list(IncidenteSeguranca.STATUS_CHOICES),
            "aud_status_choices": [("", _("Todos"))] + list(AuditoriaHSE.STATUS_CHOICES),
            "acao_status_choices": [("", _("Todos"))] + list(AcaoCorretiva.STATUS_CHOICES),
            "prev_status_choices": [("", _("Todos"))] + list(AcaoPreventiva.STATUS_CHOICES),
            "plano_status_choices": [
                ("", _("Todos")),
                ("ativos", _("Ativos")),
                ("inativos", _("Inativos")),
                ("vencidos", _("Vencidos")),
            ],
            "janela_dias_choices": [
                (0, _("Total")),
                (7, _("Últimos 7 dias")),
                (30, _("Últimos 30 dias")),
                (90, _("Últimos 90 dias")),
            ],
            "today": hoje,
            "filtros": filtros_dashboard,
        },
    )


@login_required
@admin_required
def gestao_compliance_dashboard_export_csv(request):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro

    hoje = timezone.localdate()
    filtros = normalizar_filtros_dashboard_compliance(request.GET)
    dashboard_eficacia = construir_dashboard_eficacia_compliance(
        empresa=empresa,
        referencia=hoje,
        projeto_id=filtros["drill_projeto"],
        responsavel_id=filtros["drill_responsavel"],
        janela_dias=filtros["janela_dias"] or None,
    )
    csv_bytes = gerar_dashboard_compliance_csv_bytes(
        empresa=empresa,
        dashboard_eficacia=dashboard_eficacia,
        filtros=filtros,
    )

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="dashboard_compliance.csv"'
    response.write(csv_bytes)
    return response


@login_required
@admin_required
def gestao_checklist_hse_create(request):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    form = ChecklistHSEForm(request.POST or None, empresa=empresa)
    if request.method == "POST" and form.is_valid():
        item = form.save(commit=False)
        item.empresa = empresa
        item.save()
        messages.success(request, _("Checklist HSE criada com sucesso."))
        return redirect("projetos:gestao_compliance_seguranca")
    return render(request, "projetos/gestao_checklist_hse_form.html", {"form": form, "is_create": True})


@login_required
@admin_required
def gestao_checklist_hse_update(request, pk):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    item = ChecklistHSE.objects.filter(empresa=empresa, pk=pk).first()
    if not item:
        messages.error(request, _("Checklist HSE não encontrada."))
        return redirect("projetos:gestao_compliance_seguranca")
    form = ChecklistHSEForm(request.POST or None, instance=item, empresa=empresa)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, _("Checklist HSE atualizada com sucesso."))
        return redirect("projetos:gestao_compliance_seguranca")
    return render(request, "projetos/gestao_checklist_hse_form.html", {"form": form, "is_create": False, "item": item})


@login_required
@admin_required
def gestao_checklist_hse_delete(request, pk):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    item = ChecklistHSE.objects.filter(empresa=empresa, pk=pk).first()
    if not item:
        messages.error(request, _("Checklist HSE não encontrada."))
        return redirect("projetos:gestao_compliance_seguranca")
    if request.method == "POST":
        item.delete()
        messages.success(request, _("Checklist HSE apagada com sucesso."))
        return redirect("projetos:gestao_compliance_seguranca")
    return render(request, "projetos/gestao_checklist_hse_confirm_delete.html", {"item": item})


@login_required
@admin_required
def gestao_checklist_hse_evidencia_create(request, pk):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    item = ChecklistHSE.objects.filter(empresa=empresa, pk=pk).first()
    if not item:
        messages.error(request, _("Checklist HSE não encontrada."))
        return redirect("projetos:gestao_compliance_seguranca")
    return _criar_evidencia_compliance_para_origem(
        request,
        origem=item,
        origem_label=_("Checklist HSE"),
        empresa=empresa,
        kwargs_origem={"checklist": item},
    )


@login_required
@admin_required
def gestao_incidente_create(request):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    form = IncidenteSegurancaForm(request.POST or None, empresa=empresa)
    if request.method == "POST" and form.is_valid():
        item = form.save(commit=False)
        item.empresa = empresa
        item.save()
        messages.success(request, _("Incidente registado com sucesso."))
        return redirect("projetos:gestao_compliance_seguranca")
    return render(request, "projetos/gestao_incidente_form.html", {"form": form, "is_create": True})


@login_required
@admin_required
def gestao_incidente_update(request, pk):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    item = IncidenteSeguranca.objects.filter(empresa=empresa, pk=pk).first()
    if not item:
        messages.error(request, _("Incidente não encontrado."))
        return redirect("projetos:gestao_compliance_seguranca")
    form = IncidenteSegurancaForm(request.POST or None, instance=item, empresa=empresa)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, _("Incidente atualizado com sucesso."))
        return redirect("projetos:gestao_compliance_seguranca")
    return render(request, "projetos/gestao_incidente_form.html", {"form": form, "is_create": False, "item": item})


@login_required
@admin_required
def gestao_incidente_delete(request, pk):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    item = IncidenteSeguranca.objects.filter(empresa=empresa, pk=pk).first()
    if not item:
        messages.error(request, _("Incidente não encontrado."))
        return redirect("projetos:gestao_compliance_seguranca")
    if request.method == "POST":
        item.delete()
        messages.success(request, _("Incidente apagado com sucesso."))
        return redirect("projetos:gestao_compliance_seguranca")
    return render(request, "projetos/gestao_incidente_confirm_delete.html", {"item": item})


@login_required
@admin_required
def gestao_incidente_evidencia_create(request, pk):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    item = IncidenteSeguranca.objects.filter(empresa=empresa, pk=pk).first()
    if not item:
        messages.error(request, _("Incidente não encontrado."))
        return redirect("projetos:gestao_compliance_seguranca")
    return _criar_evidencia_compliance_para_origem(
        request,
        origem=item,
        origem_label=_("Incidente"),
        empresa=empresa,
        kwargs_origem={"incidente": item},
    )


@login_required
@admin_required
def gestao_incidente_estado(request, pk, estado):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    if request.method != "POST":
        messages.error(request, _("A atualização de estado deve ser feita por formulário."))
        return redirect("projetos:gestao_compliance_seguranca")
    item = IncidenteSeguranca.objects.filter(empresa=empresa, pk=pk).first()
    if not item:
        messages.error(request, _("Incidente não encontrado."))
        return redirect("projetos:gestao_compliance_seguranca")
    if estado not in {"aberto", "investigacao", "fechado"}:
        messages.error(request, _("Estado inválido."))
        return redirect("projetos:gestao_compliance_seguranca")
    item.status = estado
    item.save(update_fields=["status", "atualizado_em"])
    messages.success(request, _("Estado do incidente atualizado."))
    return redirect("projetos:gestao_compliance_seguranca")


@login_required
@admin_required
def gestao_auditoria_hse_create(request):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    form = AuditoriaHSEForm(request.POST or None, empresa=empresa)
    if request.method == "POST" and form.is_valid():
        item = form.save(commit=False)
        item.empresa = empresa
        item.save()
        messages.success(request, _("Auditoria HSE criada com sucesso."))
        return redirect("projetos:gestao_compliance_seguranca")
    return render(request, "projetos/gestao_auditoria_hse_form.html", {"form": form, "is_create": True})


@login_required
@admin_required
def gestao_auditoria_hse_update(request, pk):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    item = AuditoriaHSE.objects.filter(empresa=empresa, pk=pk).first()
    if not item:
        messages.error(request, _("Auditoria HSE não encontrada."))
        return redirect("projetos:gestao_compliance_seguranca")
    form = AuditoriaHSEForm(request.POST or None, instance=item, empresa=empresa)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, _("Auditoria HSE atualizada com sucesso."))
        return redirect("projetos:gestao_compliance_seguranca")
    return render(request, "projetos/gestao_auditoria_hse_form.html", {"form": form, "is_create": False, "item": item})


@login_required
@admin_required
def gestao_auditoria_hse_delete(request, pk):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    item = AuditoriaHSE.objects.filter(empresa=empresa, pk=pk).first()
    if not item:
        messages.error(request, _("Auditoria HSE não encontrada."))
        return redirect("projetos:gestao_compliance_seguranca")
    if request.method == "POST":
        item.delete()
        messages.success(request, _("Auditoria HSE apagada com sucesso."))
        return redirect("projetos:gestao_compliance_seguranca")
    return render(request, "projetos/gestao_auditoria_hse_confirm_delete.html", {"item": item})


@login_required
@admin_required
def gestao_auditoria_hse_evidencia_create(request, pk):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    item = AuditoriaHSE.objects.filter(empresa=empresa, pk=pk).first()
    if not item:
        messages.error(request, _("Auditoria HSE não encontrada."))
        return redirect("projetos:gestao_compliance_seguranca")
    return _criar_evidencia_compliance_para_origem(
        request,
        origem=item,
        origem_label=_("Auditoria HSE"),
        empresa=empresa,
        kwargs_origem={"auditoria": item},
    )


@login_required
@admin_required
def gestao_plano_auditoria_hse_create(request):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    form = PlanoAuditoriaHSEForm(request.POST or None, empresa=empresa)
    if request.method == "POST" and form.is_valid():
        item = form.save(commit=False)
        item.empresa = empresa
        item.save()
        messages.success(request, _("Plano de auditoria HSE criado com sucesso."))
        return redirect("projetos:gestao_compliance_seguranca")
    return render(request, "projetos/gestao_plano_auditoria_hse_form.html", {"form": form, "is_create": True})


@login_required
@admin_required
def gestao_plano_auditoria_hse_update(request, pk):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    item = PlanoAuditoriaHSE.objects.filter(empresa=empresa, pk=pk).first()
    if not item:
        messages.error(request, _("Plano de auditoria HSE não encontrado."))
        return redirect("projetos:gestao_compliance_seguranca")
    form = PlanoAuditoriaHSEForm(request.POST or None, instance=item, empresa=empresa)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, _("Plano de auditoria HSE atualizado com sucesso."))
        return redirect("projetos:gestao_compliance_seguranca")
    return render(
        request,
        "projetos/gestao_plano_auditoria_hse_form.html",
        {"form": form, "is_create": False, "item": item},
    )


@login_required
@admin_required
def gestao_plano_auditoria_hse_delete(request, pk):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    item = PlanoAuditoriaHSE.objects.filter(empresa=empresa, pk=pk).first()
    if not item:
        messages.error(request, _("Plano de auditoria HSE não encontrado."))
        return redirect("projetos:gestao_compliance_seguranca")
    if request.method == "POST":
        item.delete()
        messages.success(request, _("Plano de auditoria HSE apagado com sucesso."))
        return redirect("projetos:gestao_compliance_seguranca")
    return render(request, "projetos/gestao_plano_auditoria_hse_confirm_delete.html", {"item": item})


@login_required
@admin_required
def gestao_plano_auditoria_hse_gerar(request):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    geradas = gerar_auditorias_recorrentes_pendentes(empresa=empresa, user=request.user)
    if geradas:
        messages.success(request, _("Foram geradas %(total)s auditorias recorrentes.") % {"total": len(geradas)})
    else:
        messages.info(request, _("Não existem auditorias recorrentes vencidas para gerar."))
    return redirect("projetos:gestao_compliance_seguranca")


@login_required
@admin_required
def gestao_acao_corretiva_create(request):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    form = AcaoCorretivaForm(request.POST or None, empresa=empresa)
    if request.method == "POST" and form.is_valid():
        item = form.save(commit=False)
        item.empresa = empresa
        if item.status != "concluida":
            item.concluida_em = None
        elif item.status == "concluida" and item.concluida_em is None:
            item.concluida_em = timezone.localdate()
        item.save()
        messages.success(request, _("Ação corretiva criada com sucesso."))
        return redirect("projetos:gestao_compliance_seguranca")
    return render(request, "projetos/gestao_acao_corretiva_form.html", {"form": form, "is_create": True})


@login_required
@admin_required
def gestao_acao_corretiva_update(request, pk):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    item = AcaoCorretiva.objects.filter(empresa=empresa, pk=pk).first()
    if not item:
        messages.error(request, _("Ação corretiva não encontrada."))
        return redirect("projetos:gestao_compliance_seguranca")
    form = AcaoCorretivaForm(request.POST or None, instance=item, empresa=empresa)
    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        if obj.status != "concluida":
            obj.concluida_em = None
        elif obj.concluida_em is None:
            obj.concluida_em = timezone.localdate()
        obj.save()
        messages.success(request, _("Ação corretiva atualizada com sucesso."))
        return redirect("projetos:gestao_compliance_seguranca")
    return render(request, "projetos/gestao_acao_corretiva_form.html", {"form": form, "is_create": False, "item": item})


@login_required
@admin_required
def gestao_acao_corretiva_delete(request, pk):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    item = AcaoCorretiva.objects.filter(empresa=empresa, pk=pk).first()
    if not item:
        messages.error(request, _("Ação corretiva não encontrada."))
        return redirect("projetos:gestao_compliance_seguranca")
    if request.method == "POST":
        item.delete()
        messages.success(request, _("Ação corretiva apagada com sucesso."))
        return redirect("projetos:gestao_compliance_seguranca")
    return render(request, "projetos/gestao_acao_corretiva_confirm_delete.html", {"item": item})


@login_required
@admin_required
def gestao_acao_corretiva_evidencia_create(request, pk):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    item = AcaoCorretiva.objects.filter(empresa=empresa, pk=pk).first()
    if not item:
        messages.error(request, _("Ação corretiva não encontrada."))
        return redirect("projetos:gestao_compliance_seguranca")
    return _criar_evidencia_compliance_para_origem(
        request,
        origem=item,
        origem_label=_("Ação corretiva"),
        empresa=empresa,
        kwargs_origem={"acao_corretiva": item},
    )


@login_required
@admin_required
def gestao_evidencia_compliance_delete(request, pk):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    evidencia = EvidenciaCompliance.objects.filter(empresa=empresa, pk=pk).first()
    if not evidencia:
        messages.error(request, _("Evidência não encontrada."))
        return redirect("projetos:gestao_compliance_seguranca")
    return _apagar_evidencia_compliance(request, evidencia=evidencia)


@login_required
@admin_required
def gestao_acao_corretiva_fecho(request, pk):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    item = AcaoCorretiva.objects.filter(empresa=empresa, pk=pk).select_related("fecho_formal").first()
    if not item:
        messages.error(request, _("Ação corretiva não encontrada."))
        return redirect("projetos:gestao_compliance_seguranca")
    instance = getattr(item, "fecho_formal", None)
    form = FechoAcaoCorretivaForm(request.POST or None, instance=instance)
    if request.method == "POST" and form.is_valid():
        registar_fecho_formal_acao_corretiva(
            acao=item,
            user=request.user,
            cleaned_data=form.cleaned_data,
        )
        messages.success(request, _("Fecho formal registado com sucesso."))
        return redirect("projetos:gestao_compliance_seguranca")
    return render(
        request,
        "projetos/gestao_acao_corretiva_fecho_form.html",
        {"form": form, "item": item, "is_create": instance is None},
    )


@login_required
@admin_required
def gestao_acao_corretiva_estado(request, pk, estado):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    if request.method != "POST":
        messages.error(request, _("A atualização de estado deve ser feita por formulário."))
        return redirect("projetos:gestao_compliance_seguranca")
    item = AcaoCorretiva.objects.filter(empresa=empresa, pk=pk).first()
    if not item:
        messages.error(request, _("Ação corretiva não encontrada."))
        return redirect("projetos:gestao_compliance_seguranca")
    if estado not in {"aberta", "em_andamento", "concluida", "cancelada"}:
        messages.error(request, _("Estado inválido."))
        return redirect("projetos:gestao_compliance_seguranca")
    if estado == "concluida":
        return redirect("projetos:gestao_acao_corretiva_fecho", pk=item.pk)
    item.status = estado
    item.concluida_em = timezone.localdate() if estado == "concluida" else None
    item.save(update_fields=["status", "concluida_em", "atualizado_em"])
    messages.success(request, _("Estado da ação corretiva atualizado."))
    return redirect("projetos:gestao_compliance_seguranca")


@login_required
@admin_required
def gestao_acao_preventiva_create(request):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    form = AcaoPreventivaForm(request.POST or None, empresa=empresa)
    if request.method == "POST" and form.is_valid():
        item = form.save(commit=False)
        item.empresa = empresa
        if item.status != "concluida":
            item.concluida_em = None
        elif item.concluida_em is None:
            item.concluida_em = timezone.localdate()
        item.save()
        messages.success(request, _("Ação preventiva criada com sucesso."))
        return redirect("projetos:gestao_compliance_seguranca")
    return render(request, "projetos/gestao_acao_preventiva_form.html", {"form": form, "is_create": True})


@login_required
@admin_required
def gestao_acao_preventiva_update(request, pk):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    item = AcaoPreventiva.objects.filter(empresa=empresa, pk=pk).first()
    if not item:
        messages.error(request, _("Ação preventiva não encontrada."))
        return redirect("projetos:gestao_compliance_seguranca")
    form = AcaoPreventivaForm(request.POST or None, instance=item, empresa=empresa)
    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        if obj.status != "concluida":
            obj.concluida_em = None
        elif obj.concluida_em is None:
            obj.concluida_em = timezone.localdate()
        obj.save()
        messages.success(request, _("Ação preventiva atualizada com sucesso."))
        return redirect("projetos:gestao_compliance_seguranca")
    return render(request, "projetos/gestao_acao_preventiva_form.html", {"form": form, "is_create": False, "item": item})


@login_required
@admin_required
def gestao_acao_preventiva_delete(request, pk):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    item = AcaoPreventiva.objects.filter(empresa=empresa, pk=pk).first()
    if not item:
        messages.error(request, _("Ação preventiva não encontrada."))
        return redirect("projetos:gestao_compliance_seguranca")
    if request.method == "POST":
        item.delete()
        messages.success(request, _("Ação preventiva apagada com sucesso."))
        return redirect("projetos:gestao_compliance_seguranca")
    return render(request, "projetos/gestao_acao_preventiva_confirm_delete.html", {"item": item})


@login_required
@admin_required
def gestao_acao_preventiva_evidencia_create(request, pk):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    item = AcaoPreventiva.objects.filter(empresa=empresa, pk=pk).first()
    if not item:
        messages.error(request, _("Ação preventiva não encontrada."))
        return redirect("projetos:gestao_compliance_seguranca")
    return _criar_evidencia_compliance_para_origem(
        request,
        origem=item,
        origem_label=_("Ação preventiva"),
        empresa=empresa,
        kwargs_origem={"acao_preventiva": item},
    )


@login_required
@admin_required
def gestao_acao_preventiva_estado(request, pk, estado):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    if request.method != "POST":
        messages.error(request, _("A atualização de estado deve ser feita por formulário."))
        return redirect("projetos:gestao_compliance_seguranca")
    item = AcaoPreventiva.objects.filter(empresa=empresa, pk=pk).first()
    if not item:
        messages.error(request, _("Ação preventiva não encontrada."))
        return redirect("projetos:gestao_compliance_seguranca")
    if estado not in {"aberta", "em_andamento", "concluida", "cancelada"}:
        messages.error(request, _("Estado inválido."))
        return redirect("projetos:gestao_compliance_seguranca")
    item.status = estado
    item.concluida_em = timezone.localdate() if estado == "concluida" else None
    item.save(update_fields=["status", "concluida_em", "atualizado_em"])
    messages.success(request, _("Estado da ação preventiva atualizado."))
    return redirect("projetos:gestao_compliance_seguranca")


@login_required
@admin_required
def gestao_notificacoes(request):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro

    pendentes_empregado = Empregados.objects.filter(empresa=empresa, aprovado=False).count()
    assiduidade_pendente = AssiduidadeRegisto.objects.filter(empresa=empresa, estado="pendente").count()
    avarias_abertas = MaquinaAvaria.objects.filter(empresa=empresa).exclude(status="resolvida").count()
    planeamentos_pendentes = PlaneamentoTurno.objects.filter(empresa=empresa, estado="planeado").count()

    fila = []
    if pendentes_empregado:
        fila.append([_("Alta"), _("Empregados pendentes"), str(pendentes_empregado)])
    if avarias_abertas:
        fila.append([_("Alta"), _("Avarias em aberto"), str(avarias_abertas)])
    if assiduidade_pendente:
        fila.append([_("Média"), _("Assiduidade pendente"), str(assiduidade_pendente)])
    if planeamentos_pendentes:
        fila.append([_("Média"), _("Planeamentos por confirmar"), str(planeamentos_pendentes)])
    estado_filtro = request.GET.get("estado", "").strip()
    prioridade_filtro = request.GET.get("prioridade", "").strip()
    notificacoes_qs = NotificacaoGestao.objects.filter(empresa=empresa)
    if estado_filtro:
        notificacoes_qs = notificacoes_qs.filter(estado=estado_filtro)
    if prioridade_filtro:
        notificacoes_qs = notificacoes_qs.filter(prioridade=prioridade_filtro)
    notificacoes = (
        notificacoes_qs
        .select_related("responsavel")
        .order_by("estado", "prazo", "-criado_em")[:20]
    )

    agora = timezone.now()
    linhas_notificacoes = []
    for n in notificacoes:
        if n.prazo is None or n.estado == "resolvida":
            sla = _("OK")
        elif n.prazo < agora:
            sla = _("Atrasado")
        elif (n.prazo - agora).total_seconds() <= 24 * 3600:
            sla = _("Em risco")
        else:
            sla = _("OK")
        linhas_notificacoes.append(
            {
                "id": n.id,
                "titulo": n.titulo,
                "prioridade": n.get_prioridade_display(),
                "estado": n.get_estado_display(),
                "responsavel": n.responsavel.nome if n.responsavel else "-",
                "prazo": n.prazo.strftime("%d/%m/%Y %H:%M") if n.prazo else "-",
                "sla": sla,
            }
        )

    return render(
        request,
        "projetos/gestao_notificacoes.html",
        {
            "titulo": _("Centro de Notificações"),
            "descricao": _("Fila operacional de pendências com prioridade."),
            "kpis": [
                {"titulo": _("Pendências críticas"), "valor": str(pendentes_empregado + avarias_abertas)},
                {"titulo": _("Pendências médias"), "valor": str(assiduidade_pendente + planeamentos_pendentes)},
            ],
            "fila": fila,
            "notificacoes": linhas_notificacoes,
            "filtros": {"estado": estado_filtro, "prioridade": prioridade_filtro},
            "estado_choices": [("", _("Todos"))] + list(NotificacaoGestao.ESTADO_CHOICES),
            "prioridade_choices": [("", _("Todos"))] + list(NotificacaoGestao.PRIORIDADE_CHOICES),
            "proximos_passos": [
                _("Adicionar atribuição por responsável e prazo limite."),
                _("Enviar alertas automáticos por email."),
                _("Criar histórico de resolução com SLA."),
            ],
        },
    )


@login_required
@admin_required
def gestao_notificacoes_export_csv(request):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro

    estado_filtro = request.GET.get("estado", "").strip()
    prioridade_filtro = request.GET.get("prioridade", "").strip()
    queryset = NotificacaoGestao.objects.filter(empresa=empresa)
    if estado_filtro:
        queryset = queryset.filter(estado=estado_filtro)
    if prioridade_filtro:
        queryset = queryset.filter(prioridade=prioridade_filtro)

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="notificacoes_gestao.csv"'
    writer = csv.writer(response)
    writer.writerow(["Título", "Tipo", "Prioridade", "Estado", "Responsável", "Prazo", "SLA", "Detalhes"])

    agora = timezone.now()
    for n in queryset.select_related("responsavel").order_by("estado", "prazo", "-criado_em"):
        if n.prazo is None or n.estado == "resolvida":
            sla = "OK"
        elif n.prazo < agora:
            sla = "Atrasado"
        elif (n.prazo - agora).total_seconds() <= 24 * 3600:
            sla = "Em risco"
        else:
            sla = "OK"
        writer.writerow(
            [
                n.titulo,
                n.tipo or "",
                n.get_prioridade_display(),
                n.get_estado_display(),
                n.responsavel.nome if n.responsavel else "",
                n.prazo.strftime("%d/%m/%Y %H:%M") if n.prazo else "",
                sla,
                n.detalhes or "",
            ]
        )
    return response


@login_required
@admin_required
def gestao_notificacao_create(request):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    form = NotificacaoGestaoForm(request.POST or None, empresa=empresa)
    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.empresa = empresa
        obj.save()
        messages.success(request, "Notificação criada com sucesso.")
        return redirect("projetos:gestao_notificacoes")
    return render(request, "projetos/gestao_notificacao_form.html", {"form": form, "is_create": True})


@login_required
@admin_required
def gestao_notificacao_update(request, pk):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    item = NotificacaoGestao.objects.filter(empresa=empresa, pk=pk).first()
    if not item:
        messages.error(request, "Notificação não encontrada.")
        return redirect("projetos:gestao_notificacoes")
    form = NotificacaoGestaoForm(request.POST or None, instance=item, empresa=empresa)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Notificação atualizada com sucesso.")
        return redirect("projetos:gestao_notificacoes")
    return render(
        request,
        "projetos/gestao_notificacao_form.html",
        {"form": form, "is_create": False, "item": item},
    )


@login_required
@admin_required
def gestao_notificacao_delete(request, pk):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    item = NotificacaoGestao.objects.filter(empresa=empresa, pk=pk).first()
    if not item:
        messages.error(request, "Notificação não encontrada.")
        return redirect("projetos:gestao_notificacoes")
    if request.method == "POST":
        item.delete()
        messages.success(request, "Notificação apagada com sucesso.")
        return redirect("projetos:gestao_notificacoes")
    return render(request, "projetos/gestao_notificacao_confirm_delete.html", {"item": item})


@login_required
@admin_required
def gestao_notificacao_estado(request, pk, estado):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    if request.method != "POST":
        messages.error(request, "A atualização da notificação deve ser feita por formulário.")
        return redirect("projetos:gestao_notificacoes")
    item = NotificacaoGestao.objects.filter(empresa=empresa, pk=pk).first()
    if not item:
        messages.error(request, "Notificação não encontrada.")
        return redirect("projetos:gestao_notificacoes")
    if estado not in {"aberta", "em_andamento", "resolvida"}:
        messages.error(request, "Estado inválido.")
        return redirect("projetos:gestao_notificacoes")
    item.estado = estado
    item.save(update_fields=["estado", "atualizado_em"])
    messages.success(request, f"Notificação atualizada para {item.get_estado_display().lower()}.")
    return redirect("projetos:gestao_notificacoes")


@login_required
@admin_required
def gestao_relatorios_executivos(request):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro

    filtros = normalizar_filtros_relatorio_executivo(request.GET)
    relatorio = _montar_relatorio_executivo(empresa=empresa, filtros=filtros)
    default_email = (empresa.responsavel_email or empresa.email or "").strip()
    form_email = RelatorioExecutivoEmailForm(
        initial={
            "assunto": _("Relatório Executivo - %(empresa)s") % {"empresa": empresa.nome},
            "destinos": default_email,
            "incluir_csv": True,
            "incluir_xlsx": True,
            "incluir_pdf": False,
        }
    )
    agendamento = _obter_ou_criar_agendamento_relatorio(empresa)
    form_agendamento = AgendamentoRelatorioExecutivoForm(instance=agendamento)
    proximo_envio = agendamento.proximo_envio_em
    if agendamento.ativo and not proximo_envio:
        proximo_envio = calcular_proximo_envio_agendado(agendamento=agendamento)
    historico_envios = []
    try:
        historico_envios = list(
            HistoricoEnvioRelatorioExecutivo.objects.filter(empresa=empresa)
            .select_related("agendamento")
            .order_by("-criado_em")[:30]
        )
    except (ProgrammingError, OperationalError):
        messages.warning(
            request,
            _("Módulo de histórico de relatórios ainda sem migração aplicada. Executa `python manage.py migrate` para ativar."),
        )

    return render(
        request,
        "projetos/gestao_relatorios_executivos.html",
        {
            "titulo": _("Relatórios Executivos"),
            "descricao": _("KPIs de alto nível para decisões de gestão, com período configurável."),
            "filtros": filtros,
            "kpis": relatorio["kpis"],
            "tendencia": relatorio["tendencia"],
            "financeiro": relatorio["financeiro"],
            "projetos_financeiro": relatorio["projetos_financeiro"],
            "rh": relatorio["rh"],
            "compliance": relatorio["compliance"],
            "form_email": form_email,
            "form_agendamento": form_agendamento,
            "agendamento": agendamento,
            "proximo_envio": proximo_envio,
            "historico_envios": historico_envios,
        },
    )


@login_required
@admin_required
def gestao_relatorios_agendamento(request):
    if request.method != "POST":
        return redirect("projetos:gestao_relatorios_executivos")

    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro

    agendamento = _obter_ou_criar_agendamento_relatorio(empresa)
    form = AgendamentoRelatorioExecutivoForm(request.POST, instance=agendamento)
    if not form.is_valid():
        for erros in form.errors.values():
            for erro in erros:
                messages.error(request, erro)
        return redirect("projetos:gestao_relatorios_executivos")

    agendamento = form.save(commit=False)
    if agendamento.ativo:
        agendamento.proximo_envio_em = calcular_proximo_envio_agendado(agendamento=agendamento)
    else:
        agendamento.proximo_envio_em = None
    agendamento.save()

    messages.success(request, _("Agendamento do relatório executivo guardado com sucesso."))
    return redirect("projetos:gestao_relatorios_executivos")


@login_required
@admin_required
def gestao_relatorios_agendamento_executar_agora(request):
    if request.method != "POST":
        return redirect("projetos:gestao_relatorios_executivos")

    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro

    agendamento = _obter_ou_criar_agendamento_relatorio(empresa)
    try:
        _executar_envio_agendado_empresa(
            empresa=empresa,
            agendamento=agendamento,
            referencia=timezone.now(),
            origem="executar_agora",
        )
    except Exception as exc:
        logger.exception("Falha ao executar envio agendado manual. empresa_id=%s", empresa.id)
        messages.error(request, _("Erro ao executar envio imediato: %(erro)s") % {"erro": str(exc)})
        return redirect("projetos:gestao_relatorios_executivos")

    if agendamento.ativo:
        agendamento.proximo_envio_em = calcular_proximo_envio_agendado(agendamento=agendamento)
    agendamento.ultimo_envio_em = timezone.now()
    agendamento.save(update_fields=["ultimo_envio_em", "proximo_envio_em", "atualizado_em"])
    messages.success(request, _("Relatório agendado enviado com sucesso."))
    return redirect("projetos:gestao_relatorios_executivos")


@login_required
@admin_required
def gestao_relatorios_export_csv(request):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    filtros = normalizar_filtros_relatorio_executivo(request.GET)
    relatorio = _montar_relatorio_executivo(empresa=empresa, filtros=filtros)
    csv_bytes = _gerar_relatorio_csv_bytes(filtros=filtros, relatorio=relatorio)

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="relatorio_executivo.csv"'
    response.write(csv_bytes)
    return response


@login_required
@admin_required
def gestao_relatorios_export_xlsx(request):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    filtros = normalizar_filtros_relatorio_executivo(request.GET)
    relatorio = _montar_relatorio_executivo(empresa=empresa, filtros=filtros)
    xlsx_bytes = _gerar_relatorio_xlsx_bytes(filtros=filtros, relatorio=relatorio)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="relatorio_executivo.xlsx"'
    response.write(xlsx_bytes)
    return response


@login_required
@admin_required
def gestao_relatorios_export_pdf(request):
    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro
    filtros = normalizar_filtros_relatorio_executivo(request.GET)
    relatorio = _montar_relatorio_executivo(empresa=empresa, filtros=filtros)
    try:
        pdf_bytes = _gerar_relatorio_pdf_bytes(empresa=empresa, filtros=filtros, relatorio=relatorio)
    except RuntimeError as exc:
        messages.error(request, str(exc))
        return redirect(construir_url_relatorio_com_filtros(filtros=filtros))

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="relatorio_executivo.pdf"'
    response.write(pdf_bytes)
    return response


@login_required
@admin_required
def gestao_relatorios_enviar_email(request):
    if request.method != "POST":
        return redirect("projetos:gestao_relatorios_executivos")

    empresa, resposta_erro = _resolver_empresa_admin(request)
    if resposta_erro:
        return resposta_erro

    filtros = normalizar_filtros_relatorio_executivo(request.GET)
    relatorio = _montar_relatorio_executivo(empresa=empresa, filtros=filtros)

    form = RelatorioExecutivoEmailForm(request.POST)
    if not form.is_valid():
        for erros in form.errors.values():
            for erro in erros:
                messages.error(request, erro)
        return redirect(construir_url_relatorio_com_filtros(filtros=filtros))

    destinos_form = form.cleaned_data.get("destinos_lista") or []
    destinos = resolver_destinos_relatorio(empresa=empresa, destinos_form=destinos_form)

    if not destinos:
        messages.error(request, _("Define pelo menos um destinatário de email."))
        return redirect(construir_url_relatorio_com_filtros(filtros=filtros))

    assunto = (form.cleaned_data.get("assunto") or "").strip()
    if not assunto:
        assunto = _("Relatório Executivo - %(empresa)s") % {"empresa": empresa.nome}
    incluir_csv = bool(form.cleaned_data.get("incluir_csv"))
    incluir_xlsx = bool(form.cleaned_data.get("incluir_xlsx"))
    incluir_pdf = bool(form.cleaned_data.get("incluir_pdf"))
    pdf_bytes = None
    if incluir_pdf:
        try:
            pdf_bytes = _gerar_relatorio_pdf_bytes(empresa=empresa, filtros=filtros, relatorio=relatorio)
        except RuntimeError as exc:
            messages.error(request, str(exc))
            return redirect(construir_url_relatorio_com_filtros(filtros=filtros))

    try:
        resultado = enviar_relatorio_executivo_email(
            empresa=empresa,
            filtros=filtros,
            relatorio=relatorio,
            assunto=assunto,
            destinos=destinos,
            incluir_csv=incluir_csv,
            incluir_xlsx=incluir_xlsx,
            incluir_pdf=incluir_pdf,
            csv_bytes=_gerar_relatorio_csv_bytes(filtros=filtros, relatorio=relatorio),
            xlsx_bytes=_gerar_relatorio_xlsx_bytes(filtros=filtros, relatorio=relatorio),
            pdf_bytes=pdf_bytes,
        )
    except Exception as exc:
        _registar_historico_envio_relatorio(
            empresa=empresa,
            origem="manual",
            status="erro",
            assunto=assunto,
            destinos=destinos,
            incluir_csv=incluir_csv,
            incluir_xlsx=incluir_xlsx,
            incluir_pdf=incluir_pdf,
            enviados=0,
            filtros=filtros,
            relatorio=relatorio,
            erro=str(exc),
        )
        messages.error(request, _("Erro ao enviar email: %(erro)s") % {"erro": str(exc)})
        return redirect(construir_url_relatorio_com_filtros(filtros=filtros))

    _registar_historico_envio_relatorio(
        empresa=empresa,
        origem="manual",
        status="sucesso",
        assunto=assunto,
        destinos=resultado.destinos,
        incluir_csv=incluir_csv,
        incluir_xlsx=incluir_xlsx,
        incluir_pdf=incluir_pdf,
        enviados=resultado.enviados,
        filtros=filtros,
        relatorio=relatorio,
    )
    if resultado.enviados:
        messages.success(request, _("Relatório enviado por email com sucesso."))
    else:
        messages.warning(request, _("O envio foi processado mas nenhum email foi confirmado como enviado."))
    return redirect(construir_url_relatorio_com_filtros(filtros=filtros))


def _executar_envio_agendado_empresa(*, empresa, agendamento, referencia=None, origem="agendado"):
    filtros = construir_filtros_periodo_agendamento(agendamento=agendamento, referencia=referencia)
    relatorio = _montar_relatorio_executivo(empresa=empresa, filtros=filtros)
    destinos_agendamento = normalizar_destinos(agendamento.destinos or "")
    destinos = resolver_destinos_relatorio(empresa=empresa, destinos_form=destinos_agendamento)
    if not destinos:
        erro = "Não existem destinatários válidos para o agendamento."
        _registar_historico_envio_relatorio(
            empresa=empresa,
            agendamento=agendamento,
            origem=origem,
            status="erro",
            assunto=_("Relatório Executivo Agendado - %(empresa)s") % {"empresa": empresa.nome},
            destinos=destinos,
            incluir_csv=bool(agendamento.incluir_csv),
            incluir_xlsx=bool(agendamento.incluir_xlsx),
            incluir_pdf=bool(agendamento.incluir_pdf),
            enviados=0,
            filtros=filtros,
            relatorio=relatorio,
            erro=erro,
        )
        raise ValueError(erro)

    assunto = _("Relatório Executivo Agendado - %(empresa)s") % {"empresa": empresa.nome}
    incluir_pdf = bool(agendamento.incluir_pdf)
    pdf_bytes = None
    if incluir_pdf:
        pdf_bytes = _gerar_relatorio_pdf_bytes(empresa=empresa, filtros=filtros, relatorio=relatorio)
    try:
        resultado = enviar_relatorio_executivo_email(
            empresa=empresa,
            filtros=filtros,
            relatorio=relatorio,
            assunto=assunto,
            destinos=destinos,
            incluir_csv=bool(agendamento.incluir_csv),
            incluir_xlsx=bool(agendamento.incluir_xlsx),
            incluir_pdf=incluir_pdf,
            csv_bytes=_gerar_relatorio_csv_bytes(filtros=filtros, relatorio=relatorio),
            xlsx_bytes=_gerar_relatorio_xlsx_bytes(filtros=filtros, relatorio=relatorio),
            pdf_bytes=pdf_bytes,
        )
    except Exception as exc:
        _registar_historico_envio_relatorio(
            empresa=empresa,
            agendamento=agendamento,
            origem=origem,
            status="erro",
            assunto=assunto,
            destinos=destinos,
            incluir_csv=bool(agendamento.incluir_csv),
            incluir_xlsx=bool(agendamento.incluir_xlsx),
            incluir_pdf=incluir_pdf,
            enviados=0,
            filtros=filtros,
            relatorio=relatorio,
            erro=str(exc),
        )
        raise

    _registar_historico_envio_relatorio(
        empresa=empresa,
        agendamento=agendamento,
        origem=origem,
        status="sucesso",
        assunto=assunto,
        destinos=destinos,
        incluir_csv=bool(agendamento.incluir_csv),
        incluir_xlsx=bool(agendamento.incluir_xlsx),
        incluir_pdf=incluir_pdf,
        enviados=resultado.enviados,
        filtros=filtros,
        relatorio=relatorio,
    )
    return resultado


def _gerar_relatorio_csv_bytes(*, filtros, relatorio):
    buffer = io.StringIO()
    writer = csv.writer(buffer)

    writer.writerow(["Periodo inicio", filtros["data_inicio"] or "-"])
    writer.writerow(["Periodo fim", filtros["data_fim"] or "-"])
    writer.writerow([])

    writer.writerow(["Metrica", "Valor"])
    for item in relatorio["kpis"]:
        writer.writerow([item["titulo"], item["valor"]])

    writer.writerow([])
    writer.writerow(["Financeiro", "Valor"])
    writer.writerow(["Despesas total (€)", f'{relatorio["financeiro"]["despesas_total"]:.2f}'])
    writer.writerow(["Despesas quantidade", relatorio["financeiro"]["despesas_qtd"]])
    writer.writerow(["Receita estimada por projeto (€)", f'{relatorio["projetos_financeiro"]["totais"]["receita_estimada"]:.2f}'])
    writer.writerow(["Margem estimada por projeto (€)", f'{relatorio["projetos_financeiro"]["totais"]["margem_estimada"]:.2f}'])

    writer.writerow([])
    writer.writerow(["Comparativo por projeto", "", "", "", "", "", ""])
    writer.writerow(
        [
            "Projeto",
            "Metros no período",
            "Registos",
            "Custo total (€)",
            "Receita estimada (€)",
            "Margem estimada (€)",
            "Custo/m",
        ]
    )
    for item in relatorio["projetos_financeiro"]["linhas"]:
        writer.writerow(
            [
                item["projeto_nome"],
                f'{item["metros"]:.2f}',
                item["registos"],
                f'{item["custo_total"]:.2f}',
                f'{item["receita_estimada"]:.2f}',
                f'{item["margem_estimada"]:.2f}',
                f'{item["custo_por_metro"]:.2f}',
            ]
        )

    writer.writerow([])
    writer.writerow(["RH", "Valor"])
    writer.writerow(["Horas presenca aprovadas", f'{relatorio["rh"]["horas_presenca"]:.2f}'])
    writer.writerow(["Horas extra aprovadas", f'{relatorio["rh"]["horas_extra"]:.2f}'])
    writer.writerow(["Horas falta aprovadas", f'{relatorio["rh"]["horas_falta"]:.2f}'])

    writer.writerow([])
    writer.writerow(["Compliance", "Valor"])
    writer.writerow(["Checklists total", relatorio["compliance"]["checklists_total"]])
    writer.writerow(["Checklists nao conformes", relatorio["compliance"]["checklists_nao_conformes"]])
    writer.writerow(["Incidentes total", relatorio["compliance"]["incidentes_total"]])
    writer.writerow(["Incidentes abertos", relatorio["compliance"]["incidentes_abertos"]])
    writer.writerow(["Auditorias total", relatorio["compliance"]["auditorias_total"]])
    writer.writerow(["Auditorias abertas", relatorio["compliance"]["auditorias_abertas"]])
    writer.writerow(["Planos auditoria total", relatorio["compliance"]["planos_total"]])
    writer.writerow(["Planos auditoria vencidos", relatorio["compliance"]["planos_vencidos"]])
    writer.writerow(["Acoes corretivas total", relatorio["compliance"]["acoes_total"]])
    writer.writerow(["Acoes corretivas abertas", relatorio["compliance"]["acoes_abertas"]])
    writer.writerow(["Acoes preventivas total", relatorio["compliance"]["preventivas_total"]])
    writer.writerow(["Acoes preventivas abertas", relatorio["compliance"]["preventivas_abertas"]])
    writer.writerow(["Acoes preventivas vencidas", relatorio["compliance"]["preventivas_vencidas"]])
    writer.writerow(["Evidencias registadas", relatorio["compliance"]["evidencias_total"]])
    writer.writerow(["Acoes com fecho formal", relatorio["compliance"]["acoes_com_fecho_formal"]])

    writer.writerow([])
    writer.writerow(["Tendencia despesas", "", ""])
    writer.writerow(["Mes", "Total (€)", "Registos"])
    for item in relatorio["tendencia"]:
        writer.writerow([item["mes"], f'{item["total"]:.2f}', item["qtd"]])

    return buffer.getvalue().encode("utf-8-sig")


def _gerar_relatorio_xlsx_bytes(*, filtros, relatorio):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"
    ws.append(["Período início", filtros["data_inicio"] or "-"])
    ws.append(["Período fim", filtros["data_fim"] or "-"])
    ws.append([])
    ws.append(["Métrica", "Valor"])
    for item in relatorio["kpis"]:
        ws.append([item["titulo"], item["valor"]])

    ws2 = wb.create_sheet("Financeiro")
    ws2.append(["Indicador", "Valor"])
    ws2.append(["Despesas total (€)", float(f'{relatorio["financeiro"]["despesas_total"]:.2f}')])
    ws2.append(["Despesas quantidade", relatorio["financeiro"]["despesas_qtd"]])
    ws2.append(["Receita estimada por projeto (€)", float(f'{relatorio["projetos_financeiro"]["totais"]["receita_estimada"]:.2f}')])
    ws2.append(["Margem estimada por projeto (€)", float(f'{relatorio["projetos_financeiro"]["totais"]["margem_estimada"]:.2f}')])

    ws3 = wb.create_sheet("RH")
    ws3.append(["Indicador", "Valor"])
    ws3.append(["Horas presenca aprovadas", float(f'{relatorio["rh"]["horas_presenca"]:.2f}')])
    ws3.append(["Horas extra aprovadas", float(f'{relatorio["rh"]["horas_extra"]:.2f}')])
    ws3.append(["Horas falta aprovadas", float(f'{relatorio["rh"]["horas_falta"]:.2f}')])

    ws4 = wb.create_sheet("Compliance")
    ws4.append(["Indicador", "Valor"])
    ws4.append(["Checklists total", relatorio["compliance"]["checklists_total"]])
    ws4.append(["Checklists nao conformes", relatorio["compliance"]["checklists_nao_conformes"]])
    ws4.append(["Incidentes total", relatorio["compliance"]["incidentes_total"]])
    ws4.append(["Incidentes abertos", relatorio["compliance"]["incidentes_abertos"]])
    ws4.append(["Auditorias total", relatorio["compliance"]["auditorias_total"]])
    ws4.append(["Auditorias abertas", relatorio["compliance"]["auditorias_abertas"]])
    ws4.append(["Planos auditoria total", relatorio["compliance"]["planos_total"]])
    ws4.append(["Planos auditoria vencidos", relatorio["compliance"]["planos_vencidos"]])
    ws4.append(["Acoes corretivas total", relatorio["compliance"]["acoes_total"]])
    ws4.append(["Acoes corretivas abertas", relatorio["compliance"]["acoes_abertas"]])
    ws4.append(["Acoes preventivas total", relatorio["compliance"]["preventivas_total"]])
    ws4.append(["Acoes preventivas abertas", relatorio["compliance"]["preventivas_abertas"]])
    ws4.append(["Acoes preventivas vencidas", relatorio["compliance"]["preventivas_vencidas"]])
    ws4.append(["Evidencias registadas", relatorio["compliance"]["evidencias_total"]])
    ws4.append(["Acoes com fecho formal", relatorio["compliance"]["acoes_com_fecho_formal"]])

    ws5 = wb.create_sheet("Tendencia")
    ws5.append(["Mes", "Total (€)", "Registos"])
    for item in relatorio["tendencia"]:
        ws5.append([item["mes"], float(f'{item["total"]:.2f}'), item["qtd"]])

    ws6 = wb.create_sheet("Projetos")
    ws6.append(["Projeto", "Metros no período", "Registos", "Custo total (€)", "Receita estimada (€)", "Margem estimada (€)", "Custo/m"])
    for item in relatorio["projetos_financeiro"]["linhas"]:
        ws6.append(
            [
                item["projeto_nome"],
                float(f'{item["metros"]:.2f}'),
                item["registos"],
                float(f'{item["custo_total"]:.2f}'),
                float(f'{item["receita_estimada"]:.2f}'),
                float(f'{item["margem_estimada"]:.2f}'),
                float(f'{item["custo_por_metro"]:.2f}'),
            ]
        )

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _gerar_relatorio_pdf_bytes(*, empresa, filtros, relatorio):
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_LEFT, TA_RIGHT
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except Exception as exc:
        raise RuntimeError(
            "Exportação PDF indisponível: instala `reportlab` no ambiente (`pip install reportlab==4.2.2`)."
        ) from exc

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=18 * mm,
        bottomMargin=14 * mm,
    )

    palette = {
        "brand": colors.HexColor("#0f172a"),
        "brand_soft": colors.HexColor("#e2e8f0"),
        "accent": colors.HexColor("#2563eb"),
        "accent_soft": colors.HexColor("#dbeafe"),
        "ok": colors.HexColor("#047857"),
        "danger": colors.HexColor("#be123c"),
        "text": colors.HexColor("#111827"),
        "muted": colors.HexColor("#475569"),
        "line": colors.HexColor("#cbd5e1"),
        "panel": colors.white,
        "panel_alt": colors.HexColor("#f8fafc"),
    }

    styles = getSampleStyleSheet()
    styles.add(
        ParagraphStyle(
            name="ExecTitle",
            parent=styles["Title"],
            fontName="Helvetica-Bold",
            fontSize=20,
            leading=24,
            textColor=palette["brand"],
            alignment=TA_LEFT,
            spaceAfter=4,
        )
    )
    styles.add(
        ParagraphStyle(
            name="ExecMeta",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=9,
            leading=12,
            textColor=palette["muted"],
        )
    )
    styles.add(
        ParagraphStyle(
            name="HeroTitle",
            parent=styles["Title"],
            fontName="Helvetica-Bold",
            fontSize=24,
            leading=28,
            textColor=colors.white,
            alignment=TA_LEFT,
        )
    )
    styles.add(
        ParagraphStyle(
            name="HeroSub",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=10,
            leading=14,
            textColor=colors.white,
            alignment=TA_LEFT,
        )
    )
    styles.add(
        ParagraphStyle(
            name="SectionTitle",
            parent=styles["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=12,
            leading=14,
            textColor=palette["brand"],
            spaceAfter=6,
            spaceBefore=4,
        )
    )
    styles.add(
        ParagraphStyle(
            name="SmallCell",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=8.5,
            leading=11,
            textColor=palette["text"],
        )
    )
    styles.add(
        ParagraphStyle(
            name="SmallCellRight",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=8.5,
            leading=11,
            textColor=palette["text"],
            alignment=TA_RIGHT,
        )
    )
    styles.add(
        ParagraphStyle(
            name="SummaryBody",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=9,
            leading=13,
            textColor=palette["text"],
        )
    )

    def fmt_money(value):
        return f"{float(value or 0):,.2f} EUR".replace(",", "X").replace(".", ",").replace("X", ".")

    def fmt_num(value):
        return f"{float(value or 0):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    def build_kpi_card(title, value):
        inner = Table(
            [
                [Paragraph(title, styles["ExecMeta"])],
                [Paragraph(f"<b>{value}</b>", styles["SectionTitle"])],
            ],
            colWidths=[40 * mm],
        )
        inner.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), palette["panel_alt"]),
                    ("BOX", (0, 0), (-1, -1), 0.75, palette["line"]),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                    ("TOPPADDING", (0, 0), (-1, -1), 7),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
                ]
            )
        )
        return inner

    story = []

    logo_path = ""
    try:
        if getattr(empresa, "logo", None) and getattr(empresa.logo, "path", ""):
            candidate = empresa.logo.path
            if candidate and os.path.exists(candidate):
                logo_path = candidate
    except Exception:
        logo_path = ""

    logo_flowable = Paragraph("<b>Sistema Furação</b>", styles["SectionTitle"])
    if logo_path:
        try:
            logo_flowable = Image(logo_path, width=28 * mm, height=28 * mm, kind="proportional")
        except Exception:
            logo_flowable = Paragraph("<b>Sistema Furação</b>", styles["SectionTitle"])

    periodo_inicio = filtros.get("data_inicio") or "-"
    periodo_fim = filtros.get("data_fim") or "-"
    gerado_em = timezone.localtime(timezone.now()).strftime("%d/%m/%Y %H:%M")

    header_meta_text = (
        f"<b>Empresa:</b> {empresa.nome}<br/>"
        f"<b>Período:</b> {periodo_inicio} até {periodo_fim}<br/>"
        f"<b>Gerado em:</b> {gerado_em}"
    )
    hero_text = (
        "Visão consolidada do desempenho operacional, financeiro e de compliance "
        "para apoio rápido à decisão."
    )
    hero = Table(
        [
            [
                logo_flowable,
                Paragraph("Relatório Executivo", styles["HeroTitle"]),
                Paragraph(header_meta_text, styles["HeroSub"]),
            ],
            [
                "",
                Paragraph(hero_text, styles["HeroSub"]),
                "",
            ],
        ],
        colWidths=[34 * mm, 96 * mm, 52 * mm],
    )
    hero.setStyle(
        TableStyle(
            [
                ("SPAN", (1, 1), (2, 1)),
                ("BACKGROUND", (0, 0), (-1, -1), palette["brand"]),
                ("BOX", (0, 0), (-1, -1), 0.75, palette["brand"]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 10),
                ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                ("TOPPADDING", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ]
        )
    )
    story.append(hero)
    story.append(Spacer(1, 8))

    margem_total = relatorio["projetos_financeiro"]["totais"]["margem_estimada"]
    nao_conformes = relatorio["compliance"]["checklists_nao_conformes"]
    incidentes_abertos = relatorio["compliance"]["incidentes_abertos"]
    auditorias_abertas = relatorio["compliance"]["auditorias_abertas"]
    planos_vencidos = relatorio["compliance"]["planos_vencidos"]
    acoes_abertas = relatorio["compliance"]["acoes_abertas"]
    preventivas_abertas = relatorio["compliance"]["preventivas_abertas"]
    preventivas_vencidas = relatorio["compliance"]["preventivas_vencidas"]
    evidencias_total = relatorio["compliance"]["evidencias_total"]
    acoes_com_fecho_formal = relatorio["compliance"]["acoes_com_fecho_formal"]
    destaque_margem = "positiva" if margem_total >= 0 else "negativa"
    resumo_executivo = (
        f"No período analisado, a empresa registou {relatorio['kpis'][0]['valor']} projetos, "
        f"{relatorio['kpis'][1]['valor']} furos e {relatorio['kpis'][2]['valor']} colaboradores considerados no consolidado. "
        f"A despesa total foi de {fmt_money(relatorio['financeiro']['despesas_total'])} e a margem estimada agregada encontra-se "
        f"{destaque_margem} em {fmt_money(margem_total)}. "
        f"Em compliance, existem {nao_conformes} checklist(s) não conforme(s), {incidentes_abertos} incidente(s), "
        f"{auditorias_abertas} auditoria(s), {planos_vencidos} plano(s) vencido(s), {acoes_abertas} ação(ões) corretiva(s) "
        f"e {preventivas_abertas} ação(ões) preventiva(s) ainda em aberto. "
        f"Foram registadas {evidencias_total} evidência(s), {preventivas_vencidas} preventiva(s) estão vencidas "
        f"e {acoes_com_fecho_formal} ação(ões) corretiva(s) já têm fecho formal."
    )
    resumo_box = Table(
        [[Paragraph("<b>Resumo Executivo</b>", styles["SectionTitle"])], [Paragraph(resumo_executivo, styles["SummaryBody"])]],
        colWidths=[182 * mm],
    )
    resumo_box.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), palette["accent_soft"]),
                ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                ("BOX", (0, 0), (-1, -1), 0.75, palette["line"]),
                ("LEFTPADDING", (0, 0), (-1, -1), 10),
                ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    story.append(resumo_box)
    story.append(Spacer(1, 8))

    kpi_cards = []
    for item in relatorio["kpis"]:
        kpi_cards.append(build_kpi_card(item["titulo"], item["valor"]))
    kpi_rows = [kpi_cards[:2], kpi_cards[2:4]]
    kpi_table = Table(kpi_rows, colWidths=[86 * mm, 86 * mm], rowHeights=[None, None], hAlign="LEFT")
    kpi_table.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"), ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 6), ("TOPPADDING", (0, 0), (-1, -1), 0), ("BOTTOMPADDING", (0, 0), (-1, -1), 6)]))
    story.append(kpi_table)
    story.append(Spacer(1, 6))

    resumo_table = Table(
        [
            [
                Paragraph("<b>Financeiro</b>", styles["SectionTitle"]),
                Paragraph("<b>RH</b>", styles["SectionTitle"]),
                Paragraph("<b>Compliance</b>", styles["SectionTitle"]),
            ],
            [
                Paragraph(
                    f"Despesas total: {fmt_money(relatorio['financeiro']['despesas_total'])}<br/>"
                    f"N.º despesas: {relatorio['financeiro']['despesas_qtd']}<br/>"
                    f"Receita estimada: {fmt_money(relatorio['projetos_financeiro']['totais']['receita_estimada'])}<br/>"
                    f"Margem estimada: {fmt_money(relatorio['projetos_financeiro']['totais']['margem_estimada'])}",
                    styles["SmallCell"],
                ),
                Paragraph(
                    f"Horas presenca: {fmt_num(relatorio['rh']['horas_presenca'])}<br/>"
                    f"Horas extra: {fmt_num(relatorio['rh']['horas_extra'])}<br/>"
                    f"Horas falta: {fmt_num(relatorio['rh']['horas_falta'])}",
                    styles["SmallCell"],
                ),
                Paragraph(
                    f"Checklists total: {relatorio['compliance']['checklists_total']}<br/>"
                    f"Não conformes: {relatorio['compliance']['checklists_nao_conformes']}<br/>"
                    f"Incidentes total: {relatorio['compliance']['incidentes_total']}<br/>"
                    f"Incidentes abertos: {relatorio['compliance']['incidentes_abertos']}<br/>"
                    f"Auditorias abertas: {relatorio['compliance']['auditorias_abertas']}<br/>"
                    f"Planos vencidos: {relatorio['compliance']['planos_vencidos']}<br/>"
                    f"Ações corretivas abertas: {relatorio['compliance']['acoes_abertas']}<br/>"
                    f"Ações preventivas abertas: {relatorio['compliance']['preventivas_abertas']}<br/>"
                    f"Preventivas vencidas: {relatorio['compliance']['preventivas_vencidas']}<br/>"
                    f"Evidências: {relatorio['compliance']['evidencias_total']}<br/>"
                    f"Fechos formais: {relatorio['compliance']['acoes_com_fecho_formal']}",
                    styles["SmallCell"],
                ),
            ],
        ],
        colWidths=[58 * mm, 58 * mm, 58 * mm],
    )
    resumo_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), palette["accent_soft"]),
                ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                ("BOX", (0, 0), (-1, -1), 0.75, palette["line"]),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, palette["line"]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    story.append(resumo_table)
    story.append(Spacer(1, 10))

    story.append(Paragraph("Comparativo por Projeto", styles["SectionTitle"]))
    project_rows = [
        [
            Paragraph("<b>Projeto</b>", styles["SmallCell"]),
            Paragraph("<b>Metros</b>", styles["SmallCellRight"]),
            Paragraph("<b>Registos</b>", styles["SmallCellRight"]),
            Paragraph("<b>Custo</b>", styles["SmallCellRight"]),
            Paragraph("<b>Receita</b>", styles["SmallCellRight"]),
            Paragraph("<b>Margem</b>", styles["SmallCellRight"]),
        ]
    ]
    for item in relatorio["projetos_financeiro"]["linhas"][:18]:
        margin_color = palette["ok"] if item["margem_estimada"] >= 0 else palette["danger"]
        margin_style = ParagraphStyle(
            name=f"MarginStyle{item['projeto_id']}",
            parent=styles["SmallCellRight"],
            textColor=margin_color,
        )
        project_rows.append(
            [
                Paragraph(item["projeto_nome"], styles["SmallCell"]),
                Paragraph(f"{fmt_num(item['metros'])} m", styles["SmallCellRight"]),
                Paragraph(str(item["registos"]), styles["SmallCellRight"]),
                Paragraph(fmt_money(item["custo_total"]), styles["SmallCellRight"]),
                Paragraph(fmt_money(item["receita_estimada"]), styles["SmallCellRight"]),
                Paragraph(fmt_money(item["margem_estimada"]), margin_style),
            ]
        )
    if len(project_rows) == 1:
        project_rows.append([Paragraph("Sem dados de projetos para o período selecionado.", styles["SmallCell"]), "", "", "", "", ""])

    projects_table = Table(project_rows, colWidths=[58 * mm, 23 * mm, 19 * mm, 28 * mm, 30 * mm, 26 * mm], repeatRows=1)
    projects_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), palette["brand"]),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, palette["panel_alt"]]),
                ("BOX", (0, 0), (-1, -1), 0.75, palette["line"]),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, palette["line"]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(projects_table)
    story.append(Spacer(1, 10))

    story.append(Paragraph("Tendência de Despesas", styles["SectionTitle"]))
    trend_rows = [
        [
            Paragraph("<b>Mês</b>", styles["SmallCell"]),
            Paragraph("<b>Total</b>", styles["SmallCellRight"]),
            Paragraph("<b>Registos</b>", styles["SmallCellRight"]),
        ]
    ]
    for item in relatorio["tendencia"]:
        trend_rows.append(
            [
                Paragraph(item["mes"], styles["SmallCell"]),
                Paragraph(fmt_money(item["total"]), styles["SmallCellRight"]),
                Paragraph(str(item["qtd"]), styles["SmallCellRight"]),
            ]
        )
    if len(trend_rows) == 1:
        trend_rows.append([Paragraph("Sem dados disponíveis.", styles["SmallCell"]), "", ""])

    trend_table = Table(trend_rows, colWidths=[52 * mm, 38 * mm, 28 * mm], hAlign="LEFT", repeatRows=1)
    trend_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), palette["accent"]),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, palette["panel_alt"]]),
                ("BOX", (0, 0), (-1, -1), 0.75, palette["line"]),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, palette["line"]),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(trend_table)

    def draw_page(canvas, doc_obj):
        canvas.saveState()
        canvas.setStrokeColor(palette["line"])
        canvas.setFillColor(palette["muted"])
        canvas.setFont("Helvetica", 8)
        canvas.line(doc_obj.leftMargin, 10 * mm, A4[0] - doc_obj.rightMargin, 10 * mm)
        canvas.drawString(doc_obj.leftMargin, 6 * mm, f"Sistema Furação | {empresa.nome}")
        canvas.drawRightString(A4[0] - doc_obj.rightMargin, 6 * mm, f"Página {canvas.getPageNumber()}")
        canvas.restoreState()

    doc.build(story, onFirstPage=draw_page, onLaterPages=draw_page)
    return output.getvalue()


def _montar_relatorio_executivo(*, empresa, filtros):
    despesas_qs = Despesa.objects.filter(empresa=empresa)
    assiduidade_qs = AssiduidadeRegisto.objects.filter(empresa=empresa, estado="aprovado")
    checklists_qs = ChecklistHSE.objects.filter(empresa=empresa)
    incidentes_qs = IncidenteSeguranca.objects.filter(empresa=empresa)
    auditorias_qs = AuditoriaHSE.objects.filter(empresa=empresa)
    planos_qs = PlanoAuditoriaHSE.objects.filter(empresa=empresa)
    acoes_qs = AcaoCorretiva.objects.filter(empresa=empresa)
    preventivas_qs = AcaoPreventiva.objects.filter(empresa=empresa)
    registos_qs = RegistoDiarioEmpregado.objects.filter(empresa=empresa)

    data_inicio = filtros.get("data_inicio")
    data_fim = filtros.get("data_fim")
    if data_inicio:
        despesas_qs = despesas_qs.filter(data__gte=data_inicio)
        assiduidade_qs = assiduidade_qs.filter(data_inicio__gte=data_inicio)
        checklists_qs = checklists_qs.filter(data_check__gte=data_inicio)
        incidentes_qs = incidentes_qs.filter(data_incidente__gte=data_inicio)
        auditorias_qs = auditorias_qs.filter(data_auditoria__gte=data_inicio)
        planos_qs = planos_qs.filter(criado_em__date__gte=data_inicio)
        acoes_qs = acoes_qs.filter(criado_em__date__gte=data_inicio)
        preventivas_qs = preventivas_qs.filter(criado_em__date__gte=data_inicio)
        registos_qs = registos_qs.filter(data__gte=data_inicio)
    if data_fim:
        despesas_qs = despesas_qs.filter(data__lte=data_fim)
        assiduidade_qs = assiduidade_qs.filter(data_inicio__lte=data_fim)
        checklists_qs = checklists_qs.filter(data_check__lte=data_fim)
        incidentes_qs = incidentes_qs.filter(data_incidente__lte=data_fim)
        auditorias_qs = auditorias_qs.filter(data_auditoria__lte=data_fim)
        planos_qs = planos_qs.filter(criado_em__date__lte=data_fim)
        acoes_qs = acoes_qs.filter(criado_em__date__lte=data_fim)
        preventivas_qs = preventivas_qs.filter(criado_em__date__lte=data_fim)
        registos_qs = registos_qs.filter(data__lte=data_fim)

    projetos_qtd = Projeto.objects.filter(empresa=empresa).count()
    furos_qtd = Furo.objects.filter(empresa=empresa).count()
    empregados_qtd = Empregados.objects.filter(empresa=empresa).count()
    despesas_total = despesas_qs.aggregate(total=Sum("valor"))["total"] or 0

    horas_presenca = assiduidade_qs.filter(tipo="presenca").aggregate(total=Sum("horas"))["total"] or 0
    horas_extra = assiduidade_qs.filter(tipo="hora_extra").aggregate(total=Sum("horas"))["total"] or 0
    horas_falta = assiduidade_qs.filter(tipo="falta").aggregate(total=Sum("horas"))["total"] or 0

    checklists_total = 0
    checklists_nao_conformes = 0
    incidentes_total = 0
    incidentes_abertos = 0
    auditorias_total = 0
    auditorias_abertas = 0
    planos_total = 0
    planos_vencidos = 0
    acoes_total = 0
    acoes_abertas = 0
    preventivas_total = 0
    preventivas_abertas = 0
    preventivas_vencidas = 0
    evidencias_total = 0
    acoes_com_fecho_formal = 0
    hoje = timezone.localdate()
    try:
        checklists_total = checklists_qs.count()
        checklists_nao_conformes = checklists_qs.filter(status="nao_conforme").count()
        incidentes_total = incidentes_qs.count()
        incidentes_abertos = incidentes_qs.exclude(status="fechado").count()
        auditorias_total = auditorias_qs.count()
        auditorias_abertas = auditorias_qs.exclude(status="concluida").count()
        planos_total = planos_qs.count()
        planos_vencidos = planos_qs.filter(ativo=True, proxima_execucao__lt=hoje).count()
        acoes_total = acoes_qs.count()
        acoes_abertas = acoes_qs.exclude(status__in=["concluida", "cancelada"]).count()
        preventivas_total = preventivas_qs.count()
        preventivas_abertas = preventivas_qs.exclude(status__in=["concluida", "cancelada"]).count()
        preventivas_vencidas = preventivas_qs.filter(prazo__lt=hoje).exclude(status__in=["concluida", "cancelada"]).count()
        evidencias_total = EvidenciaCompliance.objects.filter(empresa=empresa).count()
        acoes_com_fecho_formal = FechoAcaoCorretiva.objects.filter(empresa=empresa).count()
    except (ProgrammingError, OperationalError):
        pass

    tendencia_qs = (
        despesas_qs
        .annotate(mes=TruncMonth("data"))
        .values("mes")
        .annotate(total=Sum("valor"), qtd=Count("id"))
        .order_by("-mes")[:6]
    )
    tendencia = [
        {
            "mes": item["mes"].strftime("%m/%Y") if item["mes"] else "-",
            "total": float(item["total"] or 0),
            "qtd": int(item["qtd"] or 0),
        }
        for item in tendencia_qs
    ]

    despesas_por_projeto = {
        item["projeto_id"]: float(item["total"] or 0)
        for item in despesas_qs.filter(projeto__isnull=False).values("projeto_id").annotate(total=Sum("valor"))
    }
    registos_por_projeto = {
        item["projeto_id"]: {
            "metros": float(item["metros"] or 0),
            "registos": int(item["qtd"] or 0),
        }
        for item in registos_qs.filter(projeto__isnull=False).values("projeto_id").annotate(
            metros=Sum("metros_furados"),
            qtd=Count("id"),
        )
    }

    projetos_financeiro_linhas = []
    total_receita_estimada = 0.0
    total_margem_estimada = 0.0
    custo_por_metro_global = float(getattr(empresa, "custo_por_metro_cliente", 0) or 0)
    for projeto in Projeto.objects.filter(empresa=empresa).order_by("nome"):
        registo_info = registos_por_projeto.get(projeto.pk, {"metros": 0.0, "registos": 0})
        metros = float(registo_info["metros"] or 0)
        registos_total = int(registo_info["registos"] or 0)
        despesa_projeto = float(despesas_por_projeto.get(projeto.pk, 0) or 0)
        outros_gastos = float(projeto.outros_valores_gastos_associados or 0)
        custo_total = round(despesa_projeto + outros_gastos, 2)
        custo_por_metro_projeto = float(projeto.custo_por_metro_cliente_override or custo_por_metro_global)
        receita_estimada = round(metros * custo_por_metro_projeto, 2)
        margem_estimada = round(receita_estimada - custo_total, 2)
        total_receita_estimada += receita_estimada
        total_margem_estimada += margem_estimada

        projetos_financeiro_linhas.append(
            {
                "projeto_id": str(projeto.pk),
                "projeto_nome": projeto.nome,
                "metros": round(metros, 2),
                "registos": registos_total,
                "custo_total": custo_total,
                "receita_estimada": receita_estimada,
                "margem_estimada": margem_estimada,
                "custo_por_metro": round((custo_total / metros), 2) if metros > 0 else 0.0,
            }
        )

    return {
        "kpis": [
            {"titulo": _("Projetos ativos"), "valor": str(projetos_qtd)},
            {"titulo": _("Furos registados"), "valor": str(furos_qtd)},
            {"titulo": _("Empregados"), "valor": str(empregados_qtd)},
            {"titulo": _("Despesa no período"), "valor": f"{despesas_total:,.2f} €"},
        ],
        "financeiro": {
            "despesas_total": float(despesas_total),
            "despesas_qtd": despesas_qs.count(),
        },
        "projetos_financeiro": {
            "linhas": projetos_financeiro_linhas,
            "totais": {
                "receita_estimada": round(total_receita_estimada, 2),
                "margem_estimada": round(total_margem_estimada, 2),
            },
        },
        "rh": {
            "horas_presenca": float(horas_presenca),
            "horas_extra": float(horas_extra),
            "horas_falta": float(horas_falta),
        },
        "compliance": {
            "checklists_total": checklists_total,
            "checklists_nao_conformes": checklists_nao_conformes,
            "incidentes_total": incidentes_total,
            "incidentes_abertos": incidentes_abertos,
            "auditorias_total": auditorias_total,
            "auditorias_abertas": auditorias_abertas,
            "planos_total": planos_total,
            "planos_vencidos": planos_vencidos,
            "acoes_total": acoes_total,
            "acoes_abertas": acoes_abertas,
            "preventivas_total": preventivas_total,
            "preventivas_abertas": preventivas_abertas,
            "preventivas_vencidas": preventivas_vencidas,
            "evidencias_total": evidencias_total,
            "acoes_com_fecho_formal": acoes_com_fecho_formal,
        },
        "tendencia": tendencia,
    }
