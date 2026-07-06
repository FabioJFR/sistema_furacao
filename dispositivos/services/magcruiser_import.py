import csv
import io
import re
from decimal import Decimal, InvalidOperation

from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Max
from django.utils.text import slugify

from dispositivos.services.ingestao import guardar_leitura_dispositivo
from dispositivos.models import LeituraBrutaDispositivo
from projetos.models import Furo, Projeto


_COL_MAP = {
    "depth": "depth",
    "md": "depth",
    "measured_depth": "depth",
    "profundidade": "depth",
    "inc": "inc",
    "inclination": "inc",
    "inclinacao": "inc",
    "dip": "inc",
    "azi": "azi",
    "azimuth": "azi",
    "azimute": "azi",
    "mag": "mag",
    "magnetismo": "mag",
    "magfield": "mag",
    "mag_field": "mag",
    "temp": "temp",
    "temperature": "temp",
    "temperatura": "temp",
    "hole": "hole_name",
    "hole_name": "hole_name",
    "drillhole": "hole_name",
    "holeid": "hole_name",
    "furo": "hole_name",
    "nome_furo": "hole_name",
}


def _to_decimal(value):
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    texto = str(value).strip().replace(",", ".")
    if not texto:
        return None
    try:
        return Decimal(texto)
    except (InvalidOperation, ValueError) as exc:
        raise ValidationError(f"Valor numérico inválido: '{value}'.") from exc


def _row_to_raw_payload(row):
    parts = [
        f"DEPTH={row['depth']}",
        f"INC={row['inc']}",
        f"AZI={row['azi']}",
    ]
    if row.get("mag") is not None:
        parts.append(f"MAG={row['mag']}")
    if row.get("temp") is not None:
        parts.append(f"TEMP={row['temp']}")
    return ";".join(parts)


def _decode_uploaded(uploaded_file):
    raw = uploaded_file.read()
    if isinstance(raw, str):
        return raw
    try:
        return raw.decode("utf-8")
    except UnicodeDecodeError:
        return raw.decode("latin-1")


def _normalize_row(row):
    normalized = {"depth": None, "inc": None, "azi": None, "mag": None, "temp": None, "hole_name": None}
    for key, value in row.items():
        canonical = _COL_MAP.get(str(key or "").strip().lower())
        if not canonical:
            continue
        if canonical == "hole_name":
            texto = str(value or "").strip()
            normalized[canonical] = texto or None
        else:
            normalized[canonical] = _to_decimal(value)

    if normalized["depth"] is None or normalized["inc"] is None or normalized["azi"] is None:
        return None
    return normalized


def _config_decimal_positivo(config, chave):
    valor = (config or {}).get(chave)
    if valor in (None, ""):
        return None
    decimal = _to_decimal(valor)
    if decimal is None or decimal <= 0:
        return None
    return decimal


def _obter_tolerancias_telemetria_empresa(empresa):
    config = getattr(empresa, "geologia_score_config", None) or {}
    tolerancias = config.get("magcruiser_tolerancias") or {}
    return {
        "max_delta_depth": _config_decimal_positivo(tolerancias, "max_delta_depth"),
        "max_delta_inc": _config_decimal_positivo(tolerancias, "max_delta_inc"),
        "max_delta_azi": _config_decimal_positivo(tolerancias, "max_delta_azi"),
    }


def _delta_azimute(valor_atual, valor_anterior):
    delta = abs(valor_atual - valor_anterior)
    return min(delta, Decimal("360") - delta)


def _validar_salto_telemetria(*, erros, prefixo, row, anterior, tolerancias):
    if not anterior or not tolerancias:
        return

    max_delta_depth = tolerancias.get("max_delta_depth")
    max_delta_inc = tolerancias.get("max_delta_inc")
    max_delta_azi = tolerancias.get("max_delta_azi")

    if max_delta_depth is not None and abs(row["depth"] - anterior["depth"]) > max_delta_depth:
        erros.append(f"{prefixo}: salto de profundidade acima da tolerância configurada ({max_delta_depth}).")
    if max_delta_inc is not None and abs(row["inc"] - anterior["inc"]) > max_delta_inc:
        erros.append(f"{prefixo}: salto de inclinação acima da tolerância configurada ({max_delta_inc}).")
    if max_delta_azi is not None and _delta_azimute(row["azi"], anterior["azi"]) > max_delta_azi:
        erros.append(f"{prefixo}: salto de azimute acima da tolerância configurada ({max_delta_azi}).")


def _validar_consistencia_lote(rows, *, tolerancias=None):
    erros = []
    profundidades_por_furo = {}
    anterior_por_furo = {}

    for index, row in enumerate(rows, start=1):
        prefixo = f"Entrada {index}"
        depth = row["depth"]
        inc = row["inc"]
        azi = row["azi"]
        chave_furo = _normalizar_nome_furo(row.get("hole_name")) or "__sessao__"

        if depth < 0:
            erros.append(f"{prefixo}: profundidade não pode ser negativa.")
        if inc < Decimal("-90") or inc > Decimal("90"):
            erros.append(f"{prefixo}: inclinação deve estar entre -90 e 90 graus.")
        if azi < 0 or azi > Decimal("360"):
            erros.append(f"{prefixo}: azimute deve estar entre 0 e 360 graus.")

        profundidades = profundidades_por_furo.setdefault(chave_furo, set())
        if depth in profundidades:
            nome_furo = row.get("hole_name") or "furo da sessão"
            erros.append(f"{prefixo}: profundidade duplicada para {nome_furo}.")
        profundidades.add(depth)
        _validar_salto_telemetria(
            erros=erros,
            prefixo=prefixo,
            row=row,
            anterior=anterior_por_furo.get(chave_furo),
            tolerancias=tolerancias,
        )
        anterior_por_furo[chave_furo] = row

    if erros:
        detalhe = " ".join(erros[:5])
        if len(erros) > 5:
            detalhe = f"{detalhe} (+{len(erros) - 5} erros adicionais)"
        raise ValidationError(f"Telemetria em lote inválida. {detalhe}")


def _parse_csv_text(texto):
    sample = texto[:2000]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";"

    reader = csv.DictReader(io.StringIO(texto), dialect=dialect)
    rows = []
    for row in reader:
        parsed = _normalize_row(row)
        if parsed:
            rows.append(parsed)
    return rows


def _parse_las_text(texto):
    rows = []
    in_ascii = False
    for line in texto.splitlines():
        clean = line.strip()
        if not clean:
            continue
        if clean.startswith("#"):
            continue
        upper = clean.upper()
        if upper.startswith("~A"):
            in_ascii = True
            continue
        if not in_ascii:
            continue
        if clean.startswith("~"):
            break

        chunks = clean.replace(",", ".").split()
        if len(chunks) < 3:
            continue
        depth = _to_decimal(chunks[0])
        inc = _to_decimal(chunks[1])
        azi = _to_decimal(chunks[2])
        mag = _to_decimal(chunks[3]) if len(chunks) > 3 else None
        rows.append({"depth": depth, "inc": inc, "azi": azi, "mag": mag, "temp": None, "hole_name": None})
    return rows


def _parse_xlsx_file(uploaded_file):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependência declarada em requirements.txt
        raise ValidationError("Importação XLSX indisponível: instala `openpyxl`.") from exc

    uploaded_file.seek(0)
    try:
        workbook = load_workbook(uploaded_file, data_only=True, read_only=True)
    except Exception as exc:
        raise ValidationError("Não foi possível ler o ficheiro XLSX.") from exc

    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if not headers:
        return []

    rows = []
    for values in rows_iter:
        row = {
            header: value
            for header, value in zip(headers, values)
            if header is not None
        }
        parsed = _normalize_row(row)
        if parsed:
            rows.append(parsed)
    return rows


def _normalizar_nome_furo(nome):
    return slugify(str(nome or "").strip().lower())


def _normalizar_blocos_numericos_nome(nome_normalizado):
    partes = nome_normalizado.split("-")
    normalizadas = []
    for parte in partes:
        if parte.isdigit():
            normalizadas.append(str(int(parte)))
        else:
            normalizadas.append(parte)
    return "-".join(normalizadas)


def _chaves_reconciliacao_nome_furo(nome):
    normalizado = _normalizar_nome_furo(nome)
    if not normalizado:
        return set()
    sem_zeros = _normalizar_blocos_numericos_nome(normalizado)
    return {
        normalizado,
        sem_zeros,
        re.sub(r"[^a-z0-9]", "", normalizado),
        re.sub(r"[^a-z0-9]", "", sem_zeros),
    }


def _encontrar_furo_por_nome_reconciliado(*, empresa, nome_furo):
    chaves_entrada = _chaves_reconciliacao_nome_furo(nome_furo)
    if not chaves_entrada:
        return None

    for furo in Furo.objects.filter(empresa=empresa).order_by("data", "nome"):
        if chaves_entrada & _chaves_reconciliacao_nome_furo(furo.nome):
            return furo
    return None


def _obter_projeto_importacao(empresa):
    projeto = (
        Projeto.objects.filter(empresa=empresa, nome__iexact="Importação MagCruiser")
        .order_by("criado_em")
        .first()
    )
    if projeto:
        return projeto
    return Projeto.objects.create(
        empresa=empresa,
        nome="Importação MagCruiser",
        cliente="Importação automática",
    )


def _obter_ou_criar_furo_por_nome(*, empresa, nome_furo, criar_em_falta):
    nome = str(nome_furo or "").strip()
    if not nome:
        return None, False

    furo = Furo.objects.filter(empresa=empresa, nome__iexact=nome).order_by("data").first()
    if furo:
        return furo, False

    furo = _encontrar_furo_por_nome_reconciliado(empresa=empresa, nome_furo=nome)
    if furo:
        return furo, False

    if not criar_em_falta:
        return None, False

    projeto = _obter_projeto_importacao(empresa)
    furo = Furo.objects.create(
        projeto=projeto,
        empresa=empresa,
        nome=nome,
        tipo="fundo",
        profundidade_inicial=0.0,
        profundidade_alvo_inicial=0.0,
        profundidade_alvo_atual=0.0,
        profundidade_atual=0.0,
        profundidade_maxima_atingida=0.0,
        inclinacao_planeada_inicial=0.0,
        azimute_planeado_inicial=0.0,
    )
    return furo, True


def parse_magcruiser_file(uploaded_file):
    if not uploaded_file:
        raise ValidationError("Selecione um ficheiro para importar.")

    filename = (uploaded_file.name or "").lower()

    if filename.endswith(".xlsx"):
        rows = _parse_xlsx_file(uploaded_file)
        formato = "xlsx"
    elif filename.endswith(".las"):
        texto = _decode_uploaded(uploaded_file)
        rows = _parse_las_text(texto)
        formato = "las"
    else:
        texto = _decode_uploaded(uploaded_file)
        rows = _parse_csv_text(texto)
        formato = "csv"

    if not rows:
        raise ValidationError(
            "Não foi possível extrair linhas válidas. Confirme colunas de profundidade, inclinação e azimute."
        )

    _validar_consistencia_lote(rows)

    preview = rows[:20]
    return {
        "formato": formato,
        "total_linhas": len(rows),
        "preview_rows": preview,
        "rows": rows,
        "filename": uploaded_file.name,
    }


@transaction.atomic
def gravar_importacao_magcruiser(*, sessao, rows, modo_aplicacao="all_existing"):
    if not sessao.furo_id:
        raise ValidationError("A sessão precisa estar associada a um furo para gravar medições.")
    if not rows:
        raise ValidationError("Não há linhas para gravar.")

    _validar_consistencia_lote(
        rows,
        tolerancias=_obter_tolerancias_telemetria_empresa(sessao.empresa),
    )

    modo = (modo_aplicacao or "all_existing").strip()
    if modo not in {"all_existing", "latest_existing", "all_create_missing"}:
        raise ValidationError("Modo de aplicação inválido.")

    criar_em_falta = modo == "all_create_missing"
    if modo == "latest_existing":
        latest_map = {}
        for row in rows:
            chave = _normalizar_nome_furo(row.get("hole_name")) or "__sessao__"
            latest_map[chave] = row
        rows_processar = list(latest_map.values())
    else:
        rows_processar = rows

    criadas = 0
    furos_criados = 0
    furos_sem_match = set()
    resumo_por_furo = {}
    ignoradas = 0
    ultima_sequencia = (
        LeituraBrutaDispositivo.objects.filter(sessao=sessao).aggregate(max_seq=Max("sequencia")).get("max_seq") or 0
    )
    for row in rows_processar:
        nome_furo = row.get("hole_name")
        furo_destino = None
        criado_agora = False
        chave_resumo = str(nome_furo).strip() if nome_furo else (sessao.furo.nome if sessao.furo_id else "Sem furo")
        if chave_resumo not in resumo_por_furo:
            resumo_por_furo[chave_resumo] = {"gravadas": 0, "ignoradas": 0, "criado": False}
        if nome_furo:
            furo_destino, criado_agora = _obter_ou_criar_furo_por_nome(
                empresa=sessao.empresa,
                nome_furo=nome_furo,
                criar_em_falta=criar_em_falta,
            )
            if not furo_destino:
                furos_sem_match.add(str(nome_furo).strip())
                resumo_por_furo[chave_resumo]["ignoradas"] += 1
                ignoradas += 1
                continue
        else:
            furo_destino = sessao.furo

        raw_payload = _row_to_raw_payload(row)
        guardar_leitura_dispositivo(sessao=sessao, raw_payload=raw_payload, furo=furo_destino)
        criadas += 1
        resumo_por_furo[chave_resumo]["gravadas"] += 1
        if criado_agora:
            furos_criados += 1
            resumo_por_furo[chave_resumo]["criado"] = True
        ultima_sequencia += 1

    return {
        "total_gravadas": criadas,
        "total_ignoradas": ignoradas,
        "ultima_sequencia": ultima_sequencia,
        "furos_criados": furos_criados,
        "furos_sem_match": sorted(furos_sem_match),
        "resumo_por_furo": resumo_por_furo,
    }
